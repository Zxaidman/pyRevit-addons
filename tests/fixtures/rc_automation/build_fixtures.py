# -*- coding: utf-8 -*-
"""Regenerate the sample schedule in every format the tool has to cope with.

    python tests/fixtures/rc_automation/build_fixtures.py

The CSVs are the source of truth -- they are what a schema change gets reviewed
in -- and everything else is built from them, so a format cannot quietly drift
away from the others. ``test_rc_automation.py`` asserts they all parse to the
same objects.

What gets written, and why each earns its place:

``sample_schedule.xlsx``
    The normal case, plus the INFO cover sheet that keeps project and units off
    the data sheets.
``sample_schedule.xlsm``
    Macro-enabled, which is what a firm's real template usually is.

    openpyxl cannot write one from scratch: ``Workbook().save("x.xlsm")`` writes
    an ordinary xlsx and puts an xlsm name on it. openpyxl then reads it back
    perfectly happily -- it goes by content and ignores the extension -- while
    Excel compares the declared content type against the extension and refuses
    the file outright. So the part is re-declared here, after saving. A
    macro-enabled workbook containing no macros is valid; what makes it xlsm is
    the content type, not a ``vbaProject.bin``.
The legacy ``.xls`` fixture is **not** generated. ``sample_schedule-R1.xls`` was
saved by Excel and is a genuine BIFF/OLE2 file, which nothing here can write and
no placeholder can honestly imitate. An earlier attempt at one claimed in its own
text that "Excel will refuse it", which turned out to be false -- Excel's text
import opens a text file named ``.xls`` quite happily and lays the words out in
cells. A fixture that asserts something untrue about the tool it is testing is
worse than no fixture, so it is gone and the real file took its place.
``txt_sheets/``
    One tab-separated file per sheet. Revit's own schedule export writes this,
    and a folder of them is a first-class input.
``sample_schedule.csv`` / ``sample_schedule.txt``
    Every sheet in one file, separated by ``#SHEET`` rows. One attachment to
    send and one thing to diff, and it still opens in Excel as a readable
    column.

Needs openpyxl. The extension vendors a Windows build, so on Linux or macOS::

    pip install openpyxl
"""

import io
import os
import shutil
import sys
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))

SHEETS = ("FOOTING_TYPES", "FOOTING_PLACEMENT", "FOOTING_REBAR",
          "COLUMN_TYPES", "COLUMN_PLACEMENT", "COLUMN_REBAR")

#: What the schedule calls a level, against what a model calls it. Taken from
#: the first model this was run against, where the schedule said "Foundation"
#: and "Level 1" and the model said "00 Ground Lvl." and "01 1st Floor Lvl." --
#: names no amount of matching will join, which is what this sheet is for.
LEVELS_ROWS = [
    ["Schedule", "Model"],
    ["Foundation", "00 Ground Lvl."],
    ["Level 1", "01 1st Floor Lvl."],
]

#: The cover sheet the CSVs cannot carry -- they are one table each by design.
INFO_ROWS = [
    ["PROJECT", "Riverside Tower"],
    ["UNITS", "mm"],
    ["STANDARD", "BS 8666:2020"],
    ["REVISION", "C"],
    ["", ""],
    ["Note", "Lengths in millimetres. One row per type; placement is separate."],
]


def read_csv_rows(path):
    """Rows from a CSV, quote-aware, without the ``csv`` module.

    The same hand-rolled split the toolkit uses, for the same reason: the
    Outline column is one cell full of commas.
    """
    with io.open(path, encoding="utf-8") as handle:
        text = handle.read()
    sys.path.insert(0, os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(HERE))),
        "AnonGee.extension", "lib", "py3"))
    rows = []
    for line in text.replace("\r\n", "\n").split("\n"):
        if not line.strip():
            continue
        cells, current, quoted, index = [], [], False, 0
        while index < len(line):
            char = line[index]
            if char == '"':
                if quoted and index + 1 < len(line) and line[index + 1] == '"':
                    current.append('"')
                    index += 1
                else:
                    quoted = not quoted
            elif char == "," and not quoted:
                cells.append("".join(current))
                current = []
            else:
                current.append(char)
            index += 1
        cells.append("".join(current))
        rows.append(cells)
    return rows


def strip_title_block(rows):
    """Drop the FOOTING_TYPES title block -- the INFO sheet replaces it."""
    for index, row in enumerate(rows):
        if row and row[0].strip() == "TypeMark":
            return rows[index:]
    return rows


#: The workbook part's content type, per extension. Excel checks this against
#: the file name and refuses the file when they disagree -- which is the whole
#: reason an xlsx renamed to .xlsm will not open.
WORKBOOK_CONTENT_TYPES = {
    ".xlsx": "application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet.main+xml",
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    ".xltx": "application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.template.main+xml",
    ".xltm": "application/vnd.ms-excel.template.macroEnabled.main+xml",
}


def declared_workbook_content_type(path):
    """What ``[Content_Types].xml`` says ``/xl/workbook.xml`` is."""
    archive = zipfile.ZipFile(path)
    try:
        text = archive.read("[Content_Types].xml").decode("utf-8")
    finally:
        archive.close()
    marker = 'PartName="/xl/workbook.xml"'
    index = text.find(marker)
    if index == -1:
        return None
    tail = text[index:]
    key = 'ContentType="'
    start = tail.find(key) + len(key)
    return tail[start:tail.find('"', start)]


def retype_workbook(path):
    """Re-declare the workbook part so the content type matches the extension.

    Rebuilds the archive rather than editing in place: a zip entry cannot change
    length where it sits, and a half-rewritten one is a corrupt workbook.
    """
    extension = os.path.splitext(path)[1].lower()
    wanted = WORKBOOK_CONTENT_TYPES.get(extension)
    if not wanted or declared_workbook_content_type(path) == wanted:
        return path

    temporary = path + ".rebuilding"
    source = zipfile.ZipFile(path, "r")
    try:
        target = zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED)
        try:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    for other in WORKBOOK_CONTENT_TYPES.values():
                        text = text.replace(other, wanted)
                    data = text.encode("utf-8")
                target.writestr(item, data)
        finally:
            target.close()
    finally:
        source.close()

    shutil.move(temporary, path)
    return path


def build_workbook(path):
    from openpyxl import Workbook
    book = Workbook()
    book.remove(book.active)

    for name, rows in (("INFO", INFO_ROWS), ("LEVELS", LEVELS_ROWS)):
        sheet = book.create_sheet(title=name)
        for row in rows:
            sheet.append(row)

    for name in SHEETS:
        rows = read_csv_rows(os.path.join(HERE, name + ".csv"))
        if name == "FOOTING_TYPES":
            rows = strip_title_block(rows)
        sheet = book.create_sheet(title=name)
        for row in rows:
            sheet.append(row)

    book.save(path)
    # openpyxl writes the xlsx content type whatever the file is called, and
    # Excel refuses a file whose declared type disagrees with its extension.
    return retype_workbook(path)


SHEET_MARKER = "#SHEET"


def build_single_file(path, delimiter=","):
    """Every sheet in one file, separated by ``#SHEET`` rows."""
    lines = []
    for name, rows in _all_sheets():
        lines.append(delimiter.join([SHEET_MARKER, name]))
        for row in rows:
            lines.append(delimiter.join(_quote(str(cell), delimiter)
                                        for cell in row))
        lines.append("")
    with io.open(path, "w", encoding="utf-8", newline="\n") as handle:
        handle.write(u"\n".join(lines))
    return path


def _quote(cell, delimiter):
    """Wrap a cell that contains the delimiter, so the Outline survives."""
    if delimiter in cell or '"' in cell or "\n" in cell:
        return '"' + cell.replace('"', '""') + '"'
    return cell


def _all_sheets():
    """``[(name, rows)]`` for the whole workbook, INFO first."""
    sheets = [("INFO", INFO_ROWS), ("LEVELS", LEVELS_ROWS)]
    for name in SHEETS:
        rows = read_csv_rows(os.path.join(HERE, name + ".csv"))
        if name == "FOOTING_TYPES":
            rows = strip_title_block(rows)
        sheets.append((name, rows))
    return sheets


def build_text_sheets(folder):
    if not os.path.isdir(folder):
        os.makedirs(folder)
    written = []
    rows_by_sheet = _all_sheets()

    for name, rows in rows_by_sheet:
        target = os.path.join(folder, name + ".txt")
        with io.open(target, "w", encoding="utf-8", newline="\n") as handle:
            for row in rows:
                # Tab-separated, so the Outline cell needs no quoting at all --
                # it has no tabs in it.
                handle.write(u"\t".join(str(cell) for cell in row) + u"\n")
        written.append(target)
    return written


def main():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl is needed. pip install openpyxl")
        return 1

    made = [
        build_workbook(os.path.join(HERE, "sample_schedule.xlsx")),
        build_workbook(os.path.join(HERE, "sample_schedule.xlsm")),
    ]
    made.append(build_single_file(os.path.join(HERE, "sample_schedule.csv"),
                                  ","))
    made.append(build_single_file(os.path.join(HERE, "sample_schedule.txt"),
                                  "\t"))
    made.extend(build_text_sheets(os.path.join(HERE, "txt_sheets")))

    # The legacy fixture is Excel's own sample_schedule-R1.xls, which nothing
    # here can write. Two earlier placeholders are removed on sight: one was
    # named like the real schedule, and both claimed in their own text that
    # Excel would refuse them, which is false -- Excel's text import opens a
    # text file named .xls and lays the words out in cells.
    for stale in ("not_a_workbook.xls", "sample_schedule.xls"):
        path = os.path.join(HERE, stale)
        if os.path.isfile(path):
            os.remove(path)
            print("removed", stale, "(superseded by sample_schedule-R1.xls)")

    for path in made:
        note = ""
        if path.endswith((".xlsx", ".xlsm")):
            note = "  [{0}]".format(
                declared_workbook_content_type(path).rsplit(".", 1)[-1])
        print("wrote", os.path.relpath(path, HERE) + note)
    return 0


if __name__ == "__main__":
    sys.exit(main())
