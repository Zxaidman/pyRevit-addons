# -*- coding: utf-8 -*-
"""Revit-free tests for RC Automation's workbook layer.

Run from the repository root::

    python3 -m unittest discover -s tests -v

Everything under test here reads a schedule and decides whether it can be built.
None of it imports Revit. Only ``excel_engine.read_grid`` imports openpyxl, and
it is covered by :class:`ReadGridTests`, which is skipped when openpyxl is not
importable -- the extension vendors a Windows build whose numpy will not load
everywhere, so the file layer cannot be assumed testable. Everything else takes
raw lists of lists, so every rule below is checked against hand-written rows on
any machine at all. That seam is the point.
"""

import csv
import importlib.util
import io
import os
import shutil
import sys
import zipfile
import tempfile
import unittest

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if os.path.join(_ROOT, "tests") not in sys.path:
    sys.path.insert(0, os.path.join(_ROOT, "tests"))

from _rc_loader import load as _load_rc                          # noqa: E402

_rc = _load_rc()
models = _rc.models
standards = _rc.standards
excel_engine = _rc.excel_engine
validation = _rc.validation
reconcile = _rc.reconcile
rebar_spec = _rc.rebar_spec
naming = _rc.naming

_FIXTURES = os.path.join(_ROOT, "tests", "fixtures", "rc_automation")
#: Saved by Excel, not by this repository. The genuine legacy format for the
#: refusal path, and the ground truth the generated .xlsm is held to.
_LEGACY_XLS = os.path.join(_FIXTURES, "sample_schedule-R1.xls")
_EXCEL_XLSM = os.path.join(_FIXTURES, "sample_schedule-R1.xlsm")

_BBS_STANDARD = os.path.join(
    _ROOT, "AnonGee.extension", "AnonGee.tab", "Dev.panel",
    "BBS Generator.pushbutton", "standards", "BS_8666_2020.py")


# ── helpers ────────────────────────────────────────────────────────────────

def grid(text):
    """A CSV string as the list of lists ``parse_grid`` consumes."""
    return [row for row in csv.reader(io.StringIO(text.strip("\n")))]


def fixture_grids():
    """The sample workbook, one grid per sheet."""
    grids = {}
    for name in ("FOOTING_TYPES", "FOOTING_REBAR", "FOOTING_PLACEMENT",
                 "COLUMN_TYPES", "COLUMN_REBAR", "COLUMN_PLACEMENT"):
        with io.open(os.path.join(_FIXTURES, name + ".csv"),
                     encoding="utf-8") as handle:
            grids[name] = [row for row in csv.reader(handle)]
    return grids


def _openpyxl_missing():
    try:
        import openpyxl  # noqa: F401
        return False
    except Exception:
        return True


def errors(issues):
    return [i for i in issues if i.severity == models.SEVERITY_ERROR]


def warnings(issues):
    return [i for i in issues if i.severity == models.SEVERITY_WARNING]


def messages(issues):
    return " | ".join(str(i) for i in issues)


def minimal_grids(**overrides):
    """A workbook that validates clean, so a test can break one thing at a time."""
    grids = {
        "FOOTING_TYPES": grid("""
UNITS,mm
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,900,50,75,50
"""),
        "FOOTING_REBAR": grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,15,200,21
"""),
        "COLUMN_TYPES": grid("""
TypeMark,Width,Depth,Cover
C1,400,400,40
"""),
        "COLUMN_REBAR": grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,20,8,,00
C1,Tie,10,,200,51
"""),
        "FOOTING_PLACEMENT": grid("""
Mark,TypeMark,X,Y,Level
F1-A1,F1,0,0,Foundation
"""),
        "COLUMN_PLACEMENT": grid("""
Mark,TypeMark,X,Y,BaseLevel,TopLevel
C1-A1,C1,0,0,Foundation,Level 1
"""),
    }
    grids.update(overrides)
    return grids


# ── cell coercion ──────────────────────────────────────────────────────────

class CoercionTests(unittest.TestCase):

    def test_numbers_pass_through(self):
        self.assertEqual(excel_engine.coerce_number(1200), (1200.0, None))
        self.assertEqual(excel_engine.coerce_number(1200.5), (1200.5, None))

    def test_blank_is_not_an_error(self):
        for blank in (None, "", "   "):
            self.assertEqual(excel_engine.coerce_number(blank), (None, None))

    def test_thousands_separator_and_unit_suffix(self):
        self.assertEqual(excel_engine.coerce_number("1,200")[0], 1200.0)
        self.assertEqual(excel_engine.coerce_number("1200 mm")[0], 1200.0)
        self.assertEqual(excel_engine.coerce_number(" 900 ")[0], 900.0)

    def test_text_is_reported_not_guessed(self):
        value, error = excel_engine.coerce_number("N/A")
        self.assertIsNone(value)
        self.assertIn("N/A", error)

    def test_booleans_are_not_numbers(self):
        # Excel hands back True for a checkbox cell; 1.0 would be a silent lie.
        self.assertIsNone(excel_engine.coerce_number(True)[0])

    def test_whole_bar_counts_only(self):
        self.assertEqual(excel_engine.coerce_count(12.0), (12, None))
        self.assertEqual(excel_engine.coerce_count("12"), (12, None))
        value, error = excel_engine.coerce_count(12.5)
        self.assertIsNone(value)
        self.assertIn("whole number", error)

    def test_fold_ignores_case_spaces_and_underscores(self):
        for spelling in ("Cover Top", "cover_top", "COVERTOP", "Cover-Top"):
            self.assertEqual(excel_engine.fold(spelling), "covertop")


class ShapeCodeTests(unittest.TestCase):

    def test_excel_numeric_shape_codes(self):
        # A cell holding 00 comes back as int 0 or float 0.0 depending on how it
        # was formatted; all three spellings are the same shape code.
        for raw in (0, 0.0, "00", "0"):
            self.assertEqual(standards.normalise_shape_code(raw), "00")
        for raw in (51, 51.0, "51"):
            self.assertEqual(standards.normalise_shape_code(raw), "51")

    def test_non_codes_are_rejected(self):
        for raw in (13.5, "abc", "", None, 100, True):
            self.assertIsNone(standards.normalise_shape_code(raw))

    def test_standard_diameters(self):
        self.assertTrue(standards.is_standard_diameter(16))
        self.assertTrue(standards.is_standard_diameter(16.0))
        self.assertFalse(standards.is_standard_diameter(13))
        self.assertFalse(standards.is_standard_diameter(None))


# ── the grid ───────────────────────────────────────────────────────────────

class HeaderDetectionTests(unittest.TestCase):

    def test_header_is_found_below_a_title_block(self):
        data, issues = excel_engine.parse_grid(fixture_grids())
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertIn("F1", data.footing_type_by_mark)
        # Title block is 4 rows, header is row 5, so F1 is row 6 in Excel.
        self.assertEqual(data.footing_type("F1").source_row, 6)

    def test_metadata_above_the_header_is_read(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        self.assertEqual(data.metadata.get("project"), "Riverside Tower")
        self.assertEqual(data.units, "mm")

    def test_missing_header_is_one_error_naming_the_columns(self):
        grids = minimal_grids(FOOTING_TYPES=grid("""
some,unrelated,spreadsheet
1,2,3
"""))
        _, issues = excel_engine.parse_grid(grids)
        found = [i for i in errors(issues)
                 if i.sheet == "FOOTING_TYPES" and "header" in i.message]
        self.assertEqual(len(found), 1, messages(issues))
        self.assertIn("TypeMark", found[0].message)


class SheetAndColumnTests(unittest.TestCase):

    def test_missing_sheet_is_reported(self):
        grids = minimal_grids()
        del grids["COLUMN_REBAR"]
        _, issues = excel_engine.parse_grid(grids)
        self.assertTrue(any("COLUMN_REBAR" in i.message for i in errors(issues)),
                        messages(issues))

    def test_sheet_names_tolerate_spacing_and_case(self):
        grids = minimal_grids()
        grids["footing types"] = grids.pop("FOOTING_TYPES")
        data, issues = excel_engine.parse_grid(grids)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertIn("F1", data.footing_type_by_mark)

    def test_missing_column_is_reported_once_not_per_row(self):
        grids = minimal_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Count,Spacing,ShapeCode
F1,B1,X,15,200,21
F1,B2,Y,15,200,21
F1,T1,X,,250,00
"""))
        _, issues = excel_engine.parse_grid(grids)
        missing = [i for i in errors(issues) if "Diameter" in i.message]
        self.assertEqual(len(missing), 1, messages(issues))

    def test_original_spec_column_names_still_load(self):
        # The feature specification called these FootingMark and ColumnMark.
        grids = minimal_grids(FOOTING_REBAR=grid("""
FootingMark,Layer,Direction,Dia,Nos,Spacing,Shape
F1,B1,X,16,15,200,21
"""))
        data, issues = excel_engine.parse_grid(grids)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(data.footing_rebar[0].type_mark, "F1")
        self.assertEqual(data.footing_rebar[0].diameter_mm, 16.0)
        self.assertEqual(data.footing_rebar[0].count, 15)

    def test_blank_rows_are_skipped_and_keyless_rows_warn(self):
        grids = minimal_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,15,200,21
,,,,,,
,B2,Y,16,15,200,21
"""))
        data, issues = excel_engine.parse_grid(grids)
        self.assertEqual(len(data.footing_rebar), 1)
        self.assertTrue(any("no TypeMark" in i.message for i in warnings(issues)),
                        messages(issues))

    def test_placement_is_read_in_every_mode(self):
        # Parsing is mode-independent so switching mode in the window never
        # means re-reading the file.
        for mode in models.MODES:
            data, _ = excel_engine.parse_grid(minimal_grids(), mode=mode)
            self.assertEqual(len(data.footing_placement), 1, mode)

    def test_rebar_modes_do_not_demand_placement(self):
        grids = minimal_grids()
        del grids["FOOTING_PLACEMENT"]
        del grids["COLUMN_PLACEMENT"]
        for mode in (models.MODE_REBAR_ONLY, models.MODE_RECONCILE):
            _, issues = excel_engine.parse_grid(grids, mode=mode)
            self.assertEqual(errors(issues), [], mode + ": " + messages(issues))

    def test_creating_structure_demands_placement(self):
        grids = minimal_grids()
        del grids["FOOTING_PLACEMENT"]
        _, issues = excel_engine.parse_grid(grids, mode=models.MODE_CREATE_ALL)
        self.assertTrue(any("FOOTING_PLACEMENT" in i.message
                            for i in errors(issues)), messages(issues))

    def test_unused_placement_is_noted_not_rejected(self):
        _, issues = excel_engine.parse_grid(
            minimal_grids(), mode=models.MODE_REBAR_ONLY)
        self.assertEqual(errors(issues), [], messages(issues))
        info = [i for i in issues if i.severity == models.SEVERITY_INFO]
        self.assertTrue(any("not used in this mode" in i.message for i in info),
                        messages(issues))


class UnitsTests(unittest.TestCase):

    def test_declared_millimetres(self):
        data, issues = excel_engine.parse_grid(minimal_grids())
        self.assertEqual(data.units, "mm")
        self.assertEqual(errors(issues), [], messages(issues))

    def test_other_units_are_refused(self):
        grids = minimal_grids(FOOTING_TYPES=grid("""
UNITS,m
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3,3,0.9,0.05,0.075,0.05
"""))
        _, issues = excel_engine.parse_grid(grids)
        self.assertTrue(any("UNITS" in i.message for i in errors(issues)),
                        messages(issues))

    def test_undeclared_units_warn_and_assume_mm(self):
        grids = minimal_grids(FOOTING_TYPES=grid("""
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,900,50,75,50
"""))
        data, issues = excel_engine.parse_grid(grids)
        self.assertEqual(data.units, "mm")
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertTrue(any("UNITS" in i.message for i in warnings(issues)),
                        messages(issues))


# ── layout rules and layer geometry ────────────────────────────────────────

class LayoutRuleTests(unittest.TestCase):

    def rebar(self, count=None, spacing=None):
        return models.FootingRebarRow("F1", layer="B1", direction="X",
                                      diameter_mm=16, count=count,
                                      spacing_mm=spacing)

    def test_count_and_spacing_gives_number_with_spacing(self):
        self.assertEqual(self.rebar(12, 200).layout_rule(),
                         models.LAYOUT_NUMBER_WITH_SPACING)

    def test_count_alone_gives_fixed_number(self):
        self.assertEqual(self.rebar(count=12).layout_rule(),
                         models.LAYOUT_FIXED_NUMBER)

    def test_spacing_alone_gives_maximum_spacing(self):
        self.assertEqual(self.rebar(spacing=200).layout_rule(),
                         models.LAYOUT_MAXIMUM_SPACING)

    def test_neither_has_no_rule(self):
        self.assertIsNone(self.rebar().layout_rule())


class LayerTests(unittest.TestCase):

    def test_outer_layers_are_index_zero(self):
        for layer in ("B1", "T1"):
            row = models.FootingRebarRow("F1", layer=layer)
            self.assertEqual(row.layer_index(), 0)

    def test_second_layers_are_index_one(self):
        # The creation layer turns this into a z-offset of index x diameter:
        # B2 sits one bar above B1.
        for layer in ("B2", "T2"):
            row = models.FootingRebarRow("F1", layer=layer)
            self.assertEqual(row.layer_index(), 1)

    def test_faces_are_distinguished(self):
        self.assertTrue(models.FootingRebarRow("F1", layer="B1").is_bottom)
        self.assertTrue(models.FootingRebarRow("F1", layer="T2").is_top)
        self.assertFalse(models.FootingRebarRow("F1", layer="T1").is_bottom)


# ── validation ─────────────────────────────────────────────────────────────

def validate_grids(mode=models.MODE_CREATE_ALL, **overrides):
    data, parse_issues = excel_engine.parse_grid(
        minimal_grids(**overrides), mode=mode)
    return data, parse_issues + validation.validate(data, mode=mode)


class ValidationBaselineTests(unittest.TestCase):

    def test_the_minimal_workbook_is_clean(self):
        _, issues = validate_grids()
        self.assertEqual(errors(issues), [], messages(issues))

    def test_the_sample_workbook_is_clean(self):
        data, parse_issues = excel_engine.parse_grid(fixture_grids())
        issues = parse_issues + validation.validate(data)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(len(data.footing_types), 3)
        self.assertEqual(len(data.column_types), 3)

    def test_sample_column_carries_two_main_bar_groups(self):
        # "4T20 corners + 6T16 faces" is one column, two rows -- the shape the
        # original single-row COLUMN_REBAR sheet could not express.
        data, _ = excel_engine.parse_grid(fixture_grids())
        mains = [r for r in data.column_rebar_for("C1") if r.is_main]
        self.assertEqual(sorted(r.diameter_mm for r in mains), [16.0, 20.0])

    def test_empty_workbook_is_an_error(self):
        issues = validation.validate(models.WorkbookData())
        self.assertTrue(errors(issues), messages(issues))


class TypeValidationTests(unittest.TestCase):

    def test_duplicate_type_marks_are_rejected(self):
        _, issues = validate_grids(FOOTING_TYPES=grid("""
UNITS,mm
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,900,50,75,50
F1,2400,2400,750,50,75,50
"""))
        self.assertTrue(any("already defined" in i.message for i in errors(issues)),
                        messages(issues))

    def test_duplicate_rows_survive_parsing(self):
        """The evidence has to reach the validator to be reported.

        Keying rows into a dict as they are parsed drops the second row of a
        duplicated mark, and the duplicate rule then has nothing to find. The
        list keeps every row; the lookup keeps the first.
        """
        data, _ = excel_engine.parse_grid(minimal_grids(FOOTING_TYPES=grid("""
UNITS,mm
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,900,50,75,50
F1,2400,2400,750,50,75,50
""")))
        self.assertEqual(len(data.footing_types), 2)
        self.assertEqual(len(data.footing_type_by_mark), 1)
        self.assertEqual(data.footing_type("F1").length_mm, 3000.0)

    def test_cover_thicker_than_the_footing_is_rejected(self):
        _, issues = validate_grids(FOOTING_TYPES=grid("""
UNITS,mm
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,100,50,75,50
"""))
        self.assertTrue(any("no room for bars" in i.message for i in errors(issues)),
                        messages(issues))

    def test_metres_typed_as_millimetres_are_flagged(self):
        # 3 mm long is the classic unit slip; it must not reach a transaction.
        _, issues = validate_grids(FOOTING_TYPES=grid("""
UNITS,mm
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3,3,0.9,50,75,50
"""))
        self.assertTrue(errors(issues) or warnings(issues), messages(issues))
        self.assertIn("units", messages(issues).lower())

    def test_column_cover_wider_than_the_section(self):
        _, issues = validate_grids(COLUMN_TYPES=grid("""
TypeMark,Width,Depth,Cover
C1,400,60,40
"""))
        self.assertTrue(errors(issues), messages(issues))


class RebarValidationTests(unittest.TestCase):

    def test_non_standard_diameter_is_rejected(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,13,15,200,21
"""))
        self.assertTrue(any("not a BS 8666:2020 bar size" in i.message
                            for i in errors(issues)), messages(issues))

    def test_spacing_tighter_than_the_bar_is_rejected(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,20,15,16,21
"""))
        self.assertTrue(any("overlap" in i.message for i in errors(issues)),
                        messages(issues))

    def test_neither_count_nor_spacing_is_rejected(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,,,21
"""))
        self.assertTrue(any("nothing to lay out" in i.message
                            for i in errors(issues)), messages(issues))

    def test_unknown_layer_and_direction_are_rejected(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,MIDDLE,Z,16,15,200,21
"""))
        text = messages(errors(issues))
        self.assertIn("Layer", text)
        self.assertIn("Direction", text)

    def test_one_layer_may_only_be_described_once(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,15,200,21
F1,B1,X,20,12,250,21
"""))
        self.assertTrue(any("already described" in i.message
                            for i in errors(issues)), messages(issues))

    def test_a_link_shape_is_not_a_footing_layer(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,15,200,51
"""))
        self.assertTrue(any("closed link" in i.message for i in errors(issues)),
                        messages(issues))

    def test_unsupported_shape_warns_and_does_not_block(self):
        # Shape 41 is real BS 8666 and BBS can schedule it; P0 cannot build it.
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,15,200,41
"""))
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertTrue(any("cannot build its geometry" in i.message
                            for i in warnings(issues)), messages(issues))

    def test_rebar_referring_to_a_missing_type_is_rejected(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F9,B1,X,16,15,200,21
"""))
        self.assertTrue(any("No footing type" in i.message for i in errors(issues)),
                        messages(issues))

    def test_footing_without_bottom_steel_warns(self):
        _, issues = validate_grids(FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,T1,X,16,15,200,00
"""))
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertTrue(any("no bottom reinforcement" in i.message
                            for i in warnings(issues)), messages(issues))


class ColumnRebarValidationTests(unittest.TestCase):

    def test_a_tie_must_be_a_closed_link(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,20,8,,00
C1,Tie,10,,200,00
"""))
        self.assertTrue(any("must be a closed link" in i.message
                            for i in errors(issues)), messages(issues))

    def test_a_tie_needs_a_spacing(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,20,8,,00
C1,Tie,10,,,51
"""))
        self.assertTrue(any("Spacing is required" in i.message
                            for i in errors(issues)), messages(issues))

    def test_main_bars_need_a_count(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,20,,200,00
C1,Tie,10,,200,51
"""))
        self.assertTrue(any("Count is required" in i.message
                            for i in errors(issues)), messages(issues))

    def test_half_a_confinement_zone_is_rejected(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode,SpacingEnd,ConfinementLength
C1,Main,20,8,,00,,
C1,Tie,10,,200,51,100,
"""))
        self.assertTrue(any("no zone for it to apply to" in i.message
                            for i in errors(issues)), messages(issues))

    def test_looser_end_spacing_warns(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode,SpacingEnd,ConfinementLength
C1,Main,20,8,,00,,
C1,Tie,10,,150,51,250,600
"""))
        self.assertTrue(any("tighter, not looser" in i.message
                            for i in warnings(issues)), messages(issues))

    def test_a_cage_wider_than_its_column_is_rejected(self):
        # 2 x (40 cover + 10 tie + 32 main) = 164 across a 150 mm section.
        _, issues = validate_grids(
            COLUMN_TYPES=grid("""
TypeMark,Width,Depth,Cover
C1,150,600,40
"""),
            COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,32,8,,00
C1,Tie,10,,200,51
"""))
        self.assertTrue(any("cage does not fit" in i.message
                            for i in errors(issues)), messages(issues))

    def test_two_tie_rows_for_one_column_are_ambiguous(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,20,8,,00
C1,Tie,10,,200,51
C1,Tie,10,,150,51
"""))
        self.assertTrue(any("already described" in i.message
                            for i in errors(issues)), messages(issues))

    def test_column_without_ties_warns(self):
        _, issues = validate_grids(COLUMN_REBAR=grid("""
TypeMark,BarRole,Diameter,Count,Spacing,ShapeCode
C1,Main,20,8,,00
"""))
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertTrue(any("no ties" in i.message for i in warnings(issues)),
                        messages(issues))


class IssueReportingTests(unittest.TestCase):

    def test_an_issue_names_the_cell(self):
        issue = models.Issue(models.SEVERITY_ERROR, "13 mm is not a bar size",
                             sheet="FOOTING_REBAR", row=7, column="Diameter")
        self.assertEqual(str(issue),
                         "FOOTING_REBAR row 7 · Diameter — 13 mm is not a bar size")

    def test_errors_sort_before_warnings_before_info(self):
        issues = models.sort_issues([
            models.Issue(models.SEVERITY_INFO, "counts"),
            models.Issue(models.SEVERITY_WARNING, "odd"),
            models.Issue(models.SEVERITY_ERROR, "broken"),
        ])
        self.assertEqual([i.severity for i in issues],
                         [models.SEVERITY_ERROR, models.SEVERITY_WARNING,
                          models.SEVERITY_INFO])

    def test_counts_cover_every_severity(self):
        counts = models.count_by_severity(
            [models.Issue(models.SEVERITY_ERROR, "x")])
        self.assertEqual(counts[models.SEVERITY_ERROR], 1)
        self.assertEqual(counts[models.SEVERITY_WARNING], 0)
        self.assertEqual(counts[models.SEVERITY_INFO], 0)

    def test_validation_reports_what_it_read(self):
        _, issues = validate_grids()
        info = [i for i in issues if i.severity == models.SEVERITY_INFO]
        self.assertTrue(any("footing types" in i.message for i in info),
                        messages(issues))


# ── the copy of BS 8666 must not drift ─────────────────────────────────────

class StandardsAgreementTests(unittest.TestCase):
    """``standards.py`` holds a subset of BBS Generator's BS 8666 module.

    A copy is only tolerable while something fails when the two disagree, so
    this loads the real module by path -- it cannot be imported, it lives in a
    different pushbutton -- and holds them to each other.
    """

    def bbs(self):
        spec = importlib.util.spec_from_file_location("bs8666", _BBS_STANDARD)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def test_the_standard_module_is_where_it_is_expected(self):
        self.assertTrue(os.path.isfile(_BBS_STANDARD), _BBS_STANDARD)

    def test_bar_diameters_match_the_unit_weight_table(self):
        self.assertEqual(sorted(standards.BAR_DIAMETERS_MM),
                         sorted(self.bbs().UNIT_WEIGHTS.keys()))

    def test_known_shape_codes_match(self):
        self.assertEqual(sorted(standards.KNOWN_SHAPE_CODES),
                         sorted(self.bbs().SHAPE_MAP.keys()))

    def test_supported_codes_are_a_subset_of_known_ones(self):
        self.assertTrue(
            set(standards.SUPPORTED_SHAPE_CODES) <= set(standards.KNOWN_SHAPE_CODES))
        self.assertTrue(
            set(standards.LINK_SHAPE_CODES) <= set(standards.SUPPORTED_SHAPE_CODES))

    def test_every_supported_code_has_a_description(self):
        for code in standards.SUPPORTED_SHAPE_CODES:
            self.assertIn(code, standards.SHAPE_DESCRIPTIONS)


# ── placement ──────────────────────────────────────────────────────────────

class OutlineParsingTests(unittest.TestCase):

    def test_a_polygon_reads(self):
        points, error = excel_engine.parse_outline(
            "0,0; 4500,0; 4500,3000; 2250,4200; 0,3000")
        self.assertIsNone(error)
        self.assertEqual(len(points), 5)
        self.assertEqual(points[0], (0.0, 0.0))
        self.assertEqual(points[3], (2250.0, 4200.0))

    def test_blank_is_not_an_error(self):
        self.assertEqual(excel_engine.parse_outline(""), (None, None))
        self.assertEqual(excel_engine.parse_outline(None), (None, None))

    def test_a_repeated_closing_point_is_dropped(self):
        # Writing the first point again at the end is how people close a
        # polygon; carrying it would place a zero-length edge and Revit would
        # refuse the sketch.
        points, error = excel_engine.parse_outline(
            "0,0; 3000,0; 3000,2000; 0,2000; 0,0")
        self.assertIsNone(error)
        self.assertEqual(len(points), 4)

    def test_too_few_points_is_rejected(self):
        _, error = excel_engine.parse_outline("0,0; 3000,0")
        self.assertIn("at least 3", error)

    def test_a_malformed_point_names_itself(self):
        _, error = excel_engine.parse_outline("0,0; 3000; 3000,2000")
        self.assertIn("3000", error)

    def test_a_non_numeric_point_is_rejected(self):
        _, error = excel_engine.parse_outline("0,0; a,b; 3000,2000")
        self.assertIn("numbers", error)


class PlacementValidationTests(unittest.TestCase):

    def test_the_sample_workbook_places_cleanly(self):
        data, parse_issues = excel_engine.parse_grid(fixture_grids())
        issues = parse_issues + validation.validate(data)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(len(data.footing_placement), 6)
        self.assertEqual(len(data.column_placement), 5)

    def test_the_sample_carries_a_non_rectangular_pad(self):
        # The reason footings are floors rather than family instances.
        data, _ = excel_engine.parse_grid(fixture_grids())
        shaped = [p for p in data.footing_placement if p.has_outline]
        self.assertEqual(len(shaped), 1)
        self.assertEqual(len(shaped[0].outline), 5)

    def test_grid_references_and_coordinates_both_locate_a_row(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        by_mark = dict((p.mark, p) for p in data.footing_placement)
        self.assertTrue(by_mark["F1-A1"].has_grid_reference)
        self.assertEqual(by_mark["F1-A1"].location_description(), "A-1")
        self.assertTrue(by_mark["F3-P1"].has_coordinates)
        self.assertTrue(all(p.is_locatable for p in data.footing_placement))

    def test_a_row_with_no_position_is_rejected(self):
        _, issues = validate_grids(FOOTING_PLACEMENT=grid("""
Mark,TypeMark,GridX,GridY,X,Y,Level
F1-A1,F1,,,,,Foundation
"""))
        self.assertTrue(any("No position" in i.message for i in errors(issues)),
                        messages(issues))

    def test_half_a_grid_reference_is_rejected(self):
        _, issues = validate_grids(FOOTING_PLACEMENT=grid("""
Mark,TypeMark,GridX,GridY,X,Y,Level
F1-A1,F1,A,,,,Foundation
"""))
        self.assertTrue(any("Only one grid reference" in i.message
                            for i in errors(issues)), messages(issues))

    def test_half_a_coordinate_is_rejected(self):
        _, issues = validate_grids(FOOTING_PLACEMENT=grid("""
Mark,TypeMark,GridX,GridY,X,Y,Level
F1-A1,F1,,,12500,,Foundation
"""))
        self.assertTrue(any("Only one coordinate" in i.message
                            for i in errors(issues)), messages(issues))

    def test_placing_an_undescribed_type_is_rejected(self):
        _, issues = validate_grids(FOOTING_PLACEMENT=grid("""
Mark,TypeMark,X,Y,Level
F9-A1,F9,0,0,Foundation
"""))
        self.assertTrue(any("No footing type" in i.message
                            for i in errors(issues)), messages(issues))

    def test_duplicate_instance_marks_are_rejected(self):
        _, issues = validate_grids(FOOTING_PLACEMENT=grid("""
Mark,TypeMark,X,Y,Level
F1-A1,F1,0,0,Foundation
F1-A1,F1,3000,0,Foundation
"""))
        self.assertTrue(any("already placed" in i.message
                            for i in errors(issues)), messages(issues))

    def test_a_column_between_one_level_and_itself_is_rejected(self):
        _, issues = validate_grids(COLUMN_PLACEMENT=grid("""
Mark,TypeMark,X,Y,BaseLevel,TopLevel
C1-A1,C1,0,0,Level 1,Level 1
"""))
        self.assertTrue(any("no height" in i.message for i in errors(issues)),
                        messages(issues))

    def test_a_flat_outline_is_rejected(self):
        _, issues = validate_grids(FOOTING_PLACEMENT=grid("""
Mark,TypeMark,X,Y,Level,Outline
F1-A1,F1,0,0,Foundation,"0,0; 1000,0; 2000,0"
"""))
        self.assertTrue(any("no area" in i.message for i in errors(issues)),
                        messages(issues))

    def test_a_type_that_is_never_placed_warns(self):
        _, issues = validate_grids(FOOTING_TYPES=grid("""
UNITS,mm
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,900,50,75,50
F2,2400,2400,750,50,75,50
"""), FOOTING_REBAR=grid("""
TypeMark,Layer,Direction,Diameter,Count,Spacing,ShapeCode
F1,B1,X,16,15,200,21
F2,B1,X,16,12,200,21
"""))
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertTrue(any("never placed" in i.message for i in warnings(issues)),
                        messages(issues))

    def test_placement_is_not_checked_when_nothing_is_created(self):
        # A workbook with no coordinates in it is perfectly good for putting
        # rebar into a model that is already built.
        _, issues = validate_grids(
            mode=models.MODE_REBAR_ONLY,
            FOOTING_PLACEMENT=grid("""
Mark,TypeMark,GridX,GridY,X,Y,Level
F1-A1,F1,,,,,
"""))
        self.assertEqual(errors(issues), [], messages(issues))


# ── reconciliation ─────────────────────────────────────────────────────────

class ReconciliationTests(unittest.TestCase):

    def compare(self, excel, model, **kwargs):
        return reconcile.compare("F1-A1", "Footing", excel, model, **kwargs)

    def test_agreement_produces_no_conflicts(self):
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3000.0})
        self.assertTrue(result.agrees)
        self.assertEqual(result.conflicts, [])

    def test_a_difference_is_found_and_described(self):
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        self.assertEqual(result.field_count, 1)
        self.assertEqual(result.conflicts[0].describe(),
                         "Length: 3000 in the schedule, 3200 in the model")

    def test_excel_wins_by_default(self):
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        self.assertEqual(result.source, reconcile.SOURCE_EXCEL)
        self.assertEqual(result.resolved("length_mm"), 3000.0)
        self.assertFalse(result.is_user_choice)

    def test_the_user_can_choose_the_model(self):
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        result.use_model()
        self.assertEqual(result.resolved("length_mm"), 3200.0)
        self.assertTrue(result.is_user_choice)

    def test_a_default_is_distinguishable_from_a_decision(self):
        # A report that cannot tell one from the other is not an audit trail.
        chosen = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        chosen.use_excel()
        untouched = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        self.assertEqual(chosen.source, untouched.source)
        self.assertTrue(chosen.is_user_choice)
        self.assertFalse(untouched.is_user_choice)

    def test_an_unknown_source_is_refused_not_guessed(self):
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        with self.assertRaises(ValueError):
            result.choose("Whatever")

    def test_rounding_noise_is_not_a_conflict(self):
        # The toolkit rounds to the nearest millimetre; 0.4 mm is unit
        # conversion, not a disagreement.
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3000.4})
        self.assertTrue(result.agrees)

    def test_bar_counts_are_exact(self):
        result = reconcile.compare("F1 B1", "Rebar", {"count": 15}, {"count": 14})
        self.assertEqual(result.field_count, 1)

    def test_a_field_the_model_cannot_report_is_not_a_conflict(self):
        # An unreadable parameter is "unknown", not "zero"; treating it as a
        # difference would bury the real ones.
        result = self.compare({"length_mm": 3000.0}, {})
        self.assertTrue(result.agrees)
        self.assertEqual(result.resolved("length_mm"), 3000.0)

    def test_choosing_the_model_still_falls_back_to_the_schedule(self):
        result = self.compare({"length_mm": 3000.0, "width_mm": 3000.0},
                              {"length_mm": 3200.0})
        result.use_model()
        self.assertEqual(result.resolved("length_mm"), 3200.0)
        self.assertEqual(result.resolved("width_mm"), 3000.0)

    def test_a_field_the_schedule_omits_is_not_compared(self):
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3000.0,
                                                      "thickness_mm": 900.0})
        self.assertEqual([d.field for d in result.differences], ["length_mm"])

    def test_footings_reconcile_from_workbook_objects(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        placement = data.footing_placement[0]
        result = reconcile.compare_footing(
            placement, data.footing_type(placement.type_mark),
            {"length_mm": 3200.0, "width_mm": 3000.0, "thickness_mm": 900.0})
        self.assertEqual(result.key, "F1-A1")
        self.assertEqual([d.label for d in result.conflicts], ["Length"])

    def test_columns_reconcile_from_workbook_objects(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        placement = data.column_placement[0]
        result = reconcile.compare_column(
            placement, data.column_type(placement.type_mark),
            {"width_mm": 300.0, "depth_mm": 600.0, "cover_mm": 30.0})
        self.assertEqual([d.label for d in result.conflicts], ["Cover"])

    def test_rebar_reconciles_from_a_schedule_row(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        row = data.footing_rebar[0]
        result = reconcile.compare_rebar(
            row, {"diameter_mm": 16.0, "spacing_mm": 250.0, "count": 15})
        self.assertEqual(result.key, "F1 B1")
        self.assertEqual([d.label for d in result.conflicts], ["Spacing"])

    def test_summary_counts_sides_and_decisions(self):
        agreeing = self.compare({"length_mm": 3000.0}, {"length_mm": 3000.0})
        defaulted = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        chosen = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        chosen.use_model()
        summary = reconcile.summarise([agreeing, defaulted, chosen])
        self.assertEqual(summary["compared"], 3)
        self.assertEqual(summary["matching"], 1)
        self.assertEqual(summary["differing"], 2)
        self.assertEqual(summary["using_excel"], 1)
        self.assertEqual(summary["using_model"], 1)
        self.assertEqual(summary["user_decided"], 1)

    def test_differences_are_reported_as_info_not_warnings(self):
        # Disagreeing with the model is the normal reason to run this; grading
        # it as a problem trains people to ignore the colour that means one.
        result = self.compare({"length_mm": 3000.0}, {"length_mm": 3200.0})
        found = reconcile.issues_for([result])
        self.assertEqual(len(found), 1)
        self.assertEqual(found[0].severity, models.SEVERITY_INFO)
        self.assertIn("using the excel", found[0].message.lower())


class GeometryDeferralTests(unittest.TestCase):
    """Phase 3's rules are recorded but not switched on.

    These pin what the tool does *today* -- report a geometric difference and
    change nothing -- and pin the recorded rule underneath it, so turning
    geometry changes on is a deliberate edit with tests that notice.
    """

    def geometric(self):
        return reconcile.compare("F1-A1", "Footing",
                                 {"length_mm": 3000.0}, {"length_mm": 3200.0})

    def parametric(self):
        return reconcile.compare("F1-A1", "Footing",
                                 {"cover_top_mm": 50.0}, {"cover_top_mm": 40.0})

    def test_geometry_changes_are_deferred(self):
        self.assertTrue(reconcile.GEOMETRY_CHANGES_ARE_DEFERRED)

    def test_a_dimension_difference_is_geometric(self):
        self.assertTrue(self.geometric().conflicts[0].is_geometric)
        self.assertEqual(self.geometric().actionable_conflicts, [])

    def test_a_cover_difference_is_not_geometric(self):
        # Cover is a parameter: resolving it sets a value, it does not rebuild
        # anything, so it is actionable now.
        result = self.parametric()
        self.assertFalse(result.conflicts[0].is_geometric)
        self.assertEqual(len(result.actionable_conflicts), 1)
        self.assertFalse(result.is_report_only)

    def test_every_geometric_difference_reports_only_today(self):
        for has_dependents in (True, False):
            self.assertEqual(reconcile.strategy_for(has_dependents),
                             reconcile.STRATEGY_REPORT_ONLY)
        self.assertTrue(self.geometric().is_report_only)

    def test_the_report_never_implies_a_change_that_did_not_happen(self):
        # "using the schedule" about geometry nothing rewrote is the kind of
        # sentence somebody signs a drawing off against.
        text = self.geometric().describe()
        self.assertIn("Reported only", text)
        self.assertIn("length", text)

    def test_an_agreeing_row_has_no_strategy(self):
        agreeing = reconcile.compare("F1-A1", "Footing",
                                     {"length_mm": 3000.0},
                                     {"length_mm": 3000.0})
        self.assertIsNone(agreeing.strategy())
        self.assertFalse(agreeing.is_report_only)

    def test_the_recorded_rule_is_dependency_driven(self):
        # The phase 3 decision itself: nothing depending on the element means
        # its sketch can be edited in place; dependents mean recreate and move
        # them across instead of deleting first.
        try:
            reconcile.GEOMETRY_CHANGES_ARE_DEFERRED = False
            self.assertEqual(reconcile.strategy_for(False),
                             reconcile.STRATEGY_SKETCH_EDIT)
            self.assertEqual(reconcile.strategy_for(True),
                             reconcile.STRATEGY_RECREATE_AND_REHOST)
            self.assertEqual(self.geometric().strategy(has_dependents=True),
                             reconcile.STRATEGY_RECREATE_AND_REHOST)
        finally:
            reconcile.GEOMETRY_CHANGES_ARE_DEFERRED = True

    def test_every_strategy_has_a_label(self):
        for strategy in reconcile.STRATEGIES:
            self.assertIn(strategy, reconcile.STRATEGY_LABELS)

    def test_summary_counts_what_was_only_reported(self):
        summary = reconcile.summarise([self.geometric(), self.parametric()])
        self.assertEqual(summary["differing"], 2)
        self.assertEqual(summary["report_only"], 1)


# ── bar geometry ───────────────────────────────────────────────────────────

class ScanGeometryTests(unittest.TestCase):

    RECT = [(0.0, 0.0), (3000.0, 0.0), (3000.0, 2000.0), (0.0, 2000.0)]

    def test_a_scan_line_across_a_rectangle(self):
        self.assertEqual(rebar_spec.scan_segments(self.RECT, 1000.0, "X"),
                         [(0.0, 3000.0)])
        self.assertEqual(rebar_spec.scan_segments(self.RECT, 1500.0, "Y"),
                         [(0.0, 2000.0)])

    def test_a_scan_line_outside_the_polygon_finds_nothing(self):
        self.assertEqual(rebar_spec.scan_segments(self.RECT, 5000.0, "X"), [])

    def test_a_waisted_polygon_gives_two_runs_on_one_line(self):
        # An H on its side: one scan line crosses two separate pieces of pad,
        # and each is a bar. Taking only the outermost pair would reinforce the
        # gap between them.
        h_shape = [(0.0, 0.0), (1000.0, 0.0), (1000.0, 400.0), (2000.0, 400.0),
                   (2000.0, 0.0), (3000.0, 0.0), (3000.0, 1000.0),
                   (0.0, 1000.0)]
        segments = rebar_spec.scan_segments(h_shape, 200.0, "X")
        self.assertEqual(segments, [(0.0, 1000.0), (2000.0, 3000.0)])

    def test_a_vertex_is_not_counted_twice(self):
        # A scan line through a corner must not pair a doubled crossing into a
        # zero-length bar.
        diamond = [(0.0, -1000.0), (1500.0, 0.0), (0.0, 1000.0),
                   (-1500.0, 0.0)]
        self.assertEqual(len(rebar_spec.scan_segments(diamond, 0.0, "X")), 1)

    def test_insetting_a_span(self):
        self.assertEqual(rebar_spec.inset_segment(0.0, 3000.0, 50.0),
                         (50.0, 2950.0))

    def test_a_span_that_cover_swallows_is_not_a_bar(self):
        self.assertIsNone(rebar_spec.inset_segment(0.0, 80.0, 30.0))

    def test_a_rectangle_is_centred_on_its_origin(self):
        points = rebar_spec.rectangle(3000.0, 2000.0)
        self.assertEqual(rebar_spec.bounds(points),
                         (-1500.0, -1000.0, 1500.0, 1000.0))


class BarPositionTests(unittest.TestCase):

    def test_count_and_spacing_honours_both(self):
        got = rebar_spec.bar_positions(0.0, 3000.0, count=3, spacing_mm=200.0)
        self.assertEqual(got, [1300.0, 1500.0, 1700.0])

    def test_count_alone_spreads_evenly_with_half_gaps_at_the_ends(self):
        got = rebar_spec.bar_positions(0.0, 3000.0, count=3)
        self.assertEqual(got, [500.0, 1500.0, 2500.0])

    def test_spacing_alone_fits_as_many_as_it_can(self):
        got = rebar_spec.bar_positions(0.0, 1000.0, spacing_mm=200.0)
        self.assertEqual(len(got), 6)
        self.assertAlmostEqual(got[0], 0.0)
        self.assertAlmostEqual(got[-1], 1000.0)

    def test_one_bar_sits_in_the_middle(self):
        self.assertEqual(rebar_spec.bar_positions(0.0, 1000.0, count=1),
                         [500.0])

    def test_nothing_asked_for_is_nothing_placed(self):
        self.assertEqual(rebar_spec.bar_positions(0.0, 1000.0), [])


class FootingLayerTests(unittest.TestCase):

    def footing(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        return data

    def test_a_rectangular_pad_collapses_to_one_set(self):
        data = self.footing()
        plans = rebar_spec.plan_footing(
            data.footing_type("F1"), data.footing_rebar_for("F1"))
        self.assertTrue(all(p.uniform for p in plans), [p.notes for p in plans])
        self.assertTrue(all(p.element_count == 1 for p in plans))

    def test_the_run_between_the_bends_is_the_pad_less_cover_both_ends(self):
        data = self.footing()
        plan = rebar_spec.plan_footing(
            data.footing_type("F1"), data.footing_rebar_for("F1"))[0]
        # F1 is 3000 x 3000 with 50 side cover, so the horizontal run is 2900.
        bar = plan.bars[0]
        run = abs(bar.points[-2][0] - bar.points[1][0])
        self.assertAlmostEqual(run, 2900.0)

    def test_a_top_mat_straight_bar_is_just_the_run(self):
        data = self.footing()
        top = [p for p in rebar_spec.plan_footing(
            data.footing_type("F1"), data.footing_rebar_for("F1"))
            if p.row.layer == "T1"][0]
        self.assertEqual(top.row.shape_code, "00")
        self.assertEqual(len(top.bars[0].points), 2)
        self.assertAlmostEqual(top.bars[0].length_mm, 2900.0)

    def test_a_set_carries_its_count_and_spacing(self):
        data = self.footing()
        plan = rebar_spec.plan_footing(
            data.footing_type("F1"), data.footing_rebar_for("F1"))[0]
        bar_set = plan.as_set()
        self.assertEqual(bar_set.count, 15)
        self.assertEqual(bar_set.spacing_mm, 200.0)
        self.assertEqual(bar_set.layout_rule, models.LAYOUT_NUMBER_WITH_SPACING)

    def test_a_non_rectangular_pad_becomes_a_varying_set(self):
        """One set per region of varying depth, with varying turned on.

        The five-sided pad's bars differ in length. Placed as N single bars
        that is the right steel today and nothing that follows an edit
        tomorrow; placed as one set constrained to the cover, Revit lengths
        each bar itself — which is what the ribbon's Varying Rebar Set does.
        """
        data = self.footing()
        shaped = [p for p in data.footing_placement if p.has_outline][0]
        plans = rebar_spec.plan_footing(
            data.footing_type(shaped.type_mark),
            data.footing_rebar_for(shaped.type_mark), placement=shaped)
        varying = [p for p in plans if p.varying]
        self.assertTrue(varying, "the tapered pad should vary")
        self.assertEqual(varying[0].element_count, 1)
        self.assertTrue(varying[0].as_set().varying)
        self.assertTrue(any("varying set" in n for n in varying[0].notes))

    def test_a_rectangular_pad_is_a_set_that_does_not_vary(self):
        # Orthogonal areas get their own set too, but varying stays off.
        data = self.footing()
        plan = rebar_spec.plan_footing(
            data.footing_type("F1"), data.footing_rebar_for("F1"))[0]
        self.assertTrue(plan.uniform)
        self.assertFalse(plan.varying)
        self.assertFalse(plan.as_set().varying)
        self.assertEqual(plan.element_count, 1)

    def test_layers_stack_by_the_diameter_below_them(self):
        # B2 clears the cover, then B1's whole diameter, then half its own --
        # and B1's diameter is looked up, not assumed equal to B2's.
        b1 = models.FootingRebarRow("F1", layer="B1", direction="X",
                                    diameter_mm=20.0, spacing_mm=200.0)
        b2 = models.FootingRebarRow("F1", layer="B2", direction="Y",
                                    diameter_mm=12.0, spacing_mm=200.0)
        z1 = rebar_spec.layer_elevation(b1, [b1, b2], 900.0, 50.0, 75.0)
        z2 = rebar_spec.layer_elevation(b2, [b1, b2], 900.0, 50.0, 75.0)
        self.assertAlmostEqual(z1, 75.0 + 10.0)
        self.assertAlmostEqual(z2, 75.0 + 20.0 + 6.0)

    def test_top_layers_measure_down_from_the_top_face(self):
        t1 = models.FootingRebarRow("F1", layer="T1", direction="X",
                                    diameter_mm=12.0, spacing_mm=250.0)
        z = rebar_spec.layer_elevation(t1, [t1], 900.0, 50.0, 75.0)
        self.assertAlmostEqual(z, 900.0 - 50.0 - 6.0)

    def test_cover_that_swallows_the_pad_places_nothing_and_says_why(self):
        footing = models.FootingType("F1", length_mm=400.0, width_mm=400.0,
                                     thickness_mm=900.0, cover_top_mm=50.0,
                                     cover_bottom_mm=75.0, cover_side_mm=250.0)
        row = models.FootingRebarRow("F1", layer="B1", direction="X",
                                     diameter_mm=16.0, spacing_mm=200.0)
        plan = rebar_spec.plan_footing(footing, [row])[0]
        self.assertEqual(plan.bars, [])
        self.assertTrue(plan.notes)


# The levels in the model RC Automation was first run against. Every case
# below is measured against these rather than against invented names.
REAL_LEVELS = ["00 Ground Lvl.", "01 1st Floor Lvl.", "02 2nd Floor Lvl."]


class LevelNameTests(unittest.TestCase):
    """Joining what a schedule calls a level to what the model calls it.

    The single most common reason a workbook that is perfectly correct will not
    build, and the first thing a real model proved: the schedule said
    "Foundation" and "Level 1" where the model said "00 Ground Lvl." and
    "01 1st Floor Lvl.".
    """

    def resolve(self, wanted, candidates=None):
        return naming.resolve_name(candidates or REAL_LEVELS, wanted)

    def test_an_exact_name_needs_no_explaining(self):
        matched, note = self.resolve("00 Ground Lvl.")
        self.assertEqual(matched, "00 Ground Lvl.")
        self.assertIsNone(note)

    def test_case_and_punctuation_are_forgiven(self):
        self.assertEqual(self.resolve("00 ground lvl")[0], "00 Ground Lvl.")

    def test_the_ordinal_and_the_word_level_come_off(self):
        # "00 Ground Lvl." and "Ground" are the same level; the number is the
        # model's own convention and "Lvl." says nothing about which one.
        self.assertEqual(self.resolve("Ground")[0], "00 Ground Lvl.")
        self.assertEqual(self.resolve("1st Floor")[0], "01 1st Floor Lvl.")

    def test_a_storey_number_matches_when_nothing_else_is_left(self):
        # "Level 1" is a noise word and a digit, so there is nothing to compare
        # by name at all -- but it plainly means storey 1.
        for wanted in ("Level 1", "L1", "LEVEL 1"):
            self.assertEqual(self.resolve(wanted)[0], "01 1st Floor Lvl.",
                             wanted)
        self.assertEqual(self.resolve("Level 0")[0], "00 Ground Lvl.")

    def test_a_name_the_model_does_not_have_is_refused_not_guessed(self):
        matched, note = self.resolve("Foundation")
        self.assertIsNone(matched)
        self.assertIn("Foundation", note)

    def test_a_storey_the_model_does_not_have_is_refused(self):
        self.assertIsNone(self.resolve("Level 9")[0])

    def test_two_candidates_are_named_rather_than_picked_between(self):
        """A guess that puts a foundation on the second floor is worse than a
        message naming both levels it could have meant."""
        matched, note = self.resolve(
            "Ground", ["Ground Level", "Ground Lvl", "01 1st Floor Lvl."])
        self.assertIsNone(matched)
        self.assertIn("could be", note)
        self.assertIn("Ground Level", note)

    def test_ordinals_read_every_number_in_the_name(self):
        self.assertEqual(naming.ordinal("01 1st Floor Lvl."), 1)
        self.assertEqual(naming.ordinal("02 2nd Floor Lvl."), 2)
        self.assertEqual(naming.ordinal("00 Ground Lvl."), 0)
        self.assertEqual(naming.ordinal("Level 1"), 1)
        # Numbers that disagree say nothing: "3rd of 4" is not a storey.
        self.assertIsNone(naming.ordinal("Level 3 of 4"))
        self.assertIsNone(naming.ordinal("Roof"))

    def test_significance_strips_what_a_level_is_from_which_one(self):
        self.assertEqual(naming.significant("00 Ground Lvl."), "ground")
        self.assertEqual(naming.significant("Ground Level"), "ground")
        self.assertEqual(naming.significant("Level 1"), "")

    def test_a_written_down_mapping_beats_every_guess(self):
        resolved, notes, missing = naming.build_name_map(
            REAL_LEVELS, ["Foundation"], {"Foundation": "00 Ground Lvl."})
        self.assertEqual(resolved["Foundation"], "00 Ground Lvl.")
        self.assertEqual(missing, [])
        self.assertTrue(any("mapped" in note for note in notes))

    def test_a_stale_mapping_is_reported_but_does_not_block(self):
        """Levels get renamed and the sheet does not follow.

        Blocking a run on a mapping that has gone out of date, when the name it
        was written for is sitting right there in the model, helps nobody — so
        it is said, and the matching runs anyway.
        """
        renamed = ["00 FOUNDATION LVL.", "01 GROUND LVL.", "02 1ST FLOOR LVL."]
        resolved, notes, missing = naming.build_name_map(
            renamed, ["Foundation"], {"Foundation": "00 Ground Lvl."})
        self.assertEqual(resolved["Foundation"], "00 FOUNDATION LVL.")
        self.assertEqual(missing, [])
        self.assertTrue(any("does not have" in note for note in notes), notes)
        self.assertTrue(any("instead" in note for note in notes), notes)

    def test_a_stale_mapping_for_a_name_nothing_matches_still_fails(self):
        _resolved, _notes, missing = naming.build_name_map(
            REAL_LEVELS, ["Basement"], {"Basement": "B1"})
        self.assertTrue(missing)
        self.assertIn("B1", missing[0])
        self.assertIn("Basement", missing[0])

    def test_the_renamed_model_resolves_without_help(self):
        """The model as it stood on the second run, with the levels renamed."""
        renamed = ["00 FOUNDATION LVL.", "01 GROUND LVL.", "02 1ST FLOOR LVL."]
        resolved, _notes, missing = naming.build_name_map(
            renamed, ["Foundation", "Level 1"])
        self.assertEqual(missing, [])
        self.assertEqual(resolved["Foundation"], "00 FOUNDATION LVL.")
        self.assertEqual(resolved["Level 1"], "01 GROUND LVL.")

    def test_the_real_workbook_against_the_real_model(self):
        """The case that actually happened, end to end."""
        resolved, _notes, missing = naming.build_name_map(
            REAL_LEVELS, ["Foundation", "Level 1"])
        self.assertEqual(resolved.get("Level 1"), "01 1st Floor Lvl.")
        self.assertEqual(len(missing), 1)
        self.assertIn("Foundation", missing[0])


class LevelsSheetTests(unittest.TestCase):
    """Writing down what a level means, when guessing cannot get there."""

    def with_levels(self, rows):
        grids = minimal_grids()
        grids["LEVELS"] = grid(rows)
        return grids

    def test_the_sheet_is_read(self):
        data, issues = excel_engine.parse_grid(self.with_levels("""
Schedule,Model
Foundation,00 Ground Lvl.
Level 1,01 1st Floor Lvl.
"""))
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(data.level_map["Foundation"], "00 Ground Lvl.")
        self.assertEqual(data.level_map["Level 1"], "01 1st Floor Lvl.")

    def test_a_header_row_is_not_read_as_a_mapping(self):
        data, _ = excel_engine.parse_grid(self.with_levels("""
Schedule,Model
Foundation,00 Ground Lvl.
"""))
        self.assertNotIn("Schedule", data.level_map)

    def test_the_sheet_may_be_called_a_level_map(self):
        grids = minimal_grids()
        grids["Level Map"] = grid("Foundation,00 Ground Lvl.")
        data, _ = excel_engine.parse_grid(grids)
        self.assertEqual(data.level_map["Foundation"], "00 Ground Lvl.")

    def test_no_sheet_means_no_map_and_no_complaint(self):
        data, issues = excel_engine.parse_grid(minimal_grids())
        self.assertEqual(data.level_map, {})
        self.assertEqual(errors(issues), [], messages(issues))

    def test_a_mapping_settles_what_guessing_could_not(self):
        # "Foundation" resembles nothing in the model; written down, it is
        # simply the ground level.
        data, _ = excel_engine.parse_grid(self.with_levels(
            "Foundation,00 Ground Lvl."))
        resolved, _notes, missing = naming.build_name_map(
            REAL_LEVELS, ["Foundation"], data.level_map)
        self.assertEqual(resolved["Foundation"], "00 Ground Lvl.")
        self.assertEqual(missing, [])


class GridCrossingTests(unittest.TestCase):

    def test_two_perpendicular_grids_cross_where_they_should(self):
        point, note = naming.cross_segments(
            ((0.0, 5000.0), (10000.0, 5000.0)),      # grid running east-west
            ((3000.0, 0.0), (3000.0, 9000.0)))       # grid running north-south
        self.assertIsNone(note)
        self.assertAlmostEqual(point[0], 3000.0)
        self.assertAlmostEqual(point[1], 5000.0)

    def test_grids_that_do_not_reach_each_other_still_cross(self):
        # A grid bubble stops where the drawing needed it to; a footing can sit
        # on a crossing neither drawn line reaches.
        point, note = naming.cross_segments(
            ((0.0, 0.0), (1000.0, 0.0)),
            ((5000.0, -1000.0), (5000.0, -500.0)))
        self.assertIsNone(note)
        self.assertAlmostEqual(point[0], 5000.0)
        self.assertAlmostEqual(point[1], 0.0)

    def test_parallel_grids_are_refused(self):
        point, note = naming.cross_segments(
            ((0.0, 0.0), (1000.0, 0.0)), ((0.0, 500.0), (1000.0, 500.0)))
        self.assertIsNone(point)
        self.assertIn("parallel", note)

    def test_a_skew_crossing(self):
        point, note = naming.cross_segments(
            ((0.0, 0.0), (1000.0, 1000.0)), ((0.0, 1000.0), (1000.0, 0.0)))
        self.assertIsNone(note)
        self.assertAlmostEqual(point[0], 500.0)
        self.assertAlmostEqual(point[1], 500.0)


class ScheduledExtentTests(unittest.TestCase):
    """What size the schedule says a pad is — which is what a model is held to."""

    def workbook(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        return data

    def test_a_plain_type_gives_its_length_and_width(self):
        data = self.workbook()
        self.assertEqual(
            rebar_spec.scheduled_extent_mm(data.footing_type("F1")),
            (3000.0, 3000.0))

    def test_an_outline_wins_over_the_type_rectangle(self):
        # F3's type row says 4500 x 3000; the pad placed at F3-P1 is a
        # five-sided 4500 x 4200. Holding the model to the type row would
        # reject a footing that is exactly right.
        data = self.workbook()
        shaped = [p for p in data.footing_placement if p.has_outline][0]
        footing = data.footing_type(shaped.type_mark)
        self.assertEqual(rebar_spec.scheduled_extent_mm(footing),
                         (4500.0, 3000.0))
        self.assertEqual(rebar_spec.scheduled_extent_mm(footing, shaped),
                         (4500.0, 4200.0))

    def test_a_placement_without_an_outline_falls_back_to_the_type(self):
        data = self.workbook()
        plain = [p for p in data.footing_placement if not p.has_outline][0]
        self.assertEqual(
            rebar_spec.scheduled_extent_mm(data.footing_type(plain.type_mark),
                                           plain),
            (3000.0, 3000.0))

    def test_the_outline_reaches_the_bars(self):
        # The placement has to be threaded through, or a tapered pad is
        # reinforced as the rectangle its type row describes.
        data = self.workbook()
        shaped = [p for p in data.footing_placement if p.has_outline][0]
        footing = data.footing_type(shaped.type_mark)
        rows = data.footing_rebar_for(shaped.type_mark)
        as_rectangle = rebar_spec.plan_footing(footing, rows)
        as_drawn = rebar_spec.plan_footing(footing, rows, shaped)
        # A rectangle is one plain set per layer; the drawn pad is cut into
        # regions and some of them vary.
        self.assertTrue(all(p.uniform for p in as_rectangle))
        self.assertEqual(len(as_rectangle), len(rows))
        self.assertGreater(len(as_drawn), len(rows))
        self.assertTrue(any(p.varying for p in as_drawn))


class BarShapeTests(unittest.TestCase):
    """The shape code has to build the bar, not just label it.

    Every footing bar was placed as a straight line whatever its shape code
    said. The code was carried through parsing, validation and planning as a
    label and never used to bend anything, so a bottom mat scheduled as a U-bar
    arrived in the model as shape 00 — wrong in the model, wrong in the bending
    schedule, and looking right in neither.
    """

    def points(self, shape_code, leg=600.0):
        return rebar_spec.bar_points(shape_code, 0.0, 3000.0, 500.0, 100.0,
                                     "X", leg)

    def test_a_straight_bar_is_two_points(self):
        self.assertEqual(self.points("00"),
                         [(0.0, 500.0, 100.0), (3000.0, 500.0, 100.0)])

    def test_a_u_bar_turns_up_at_both_ends(self):
        points = self.points("21")
        self.assertEqual(len(points), 4)
        self.assertEqual(points[0], (0.0, 500.0, 700.0))
        self.assertEqual(points[1], (0.0, 500.0, 100.0))
        self.assertEqual(points[2], (3000.0, 500.0, 100.0))
        self.assertEqual(points[3], (3000.0, 500.0, 700.0))

    def test_one_bend_turns_up_at_one_end(self):
        points = self.points("11")
        self.assertEqual(len(points), 3)
        self.assertEqual(points[-1], (3000.0, 500.0, 700.0))

    def test_a_bar_running_the_other_way_bends_the_same(self):
        points = rebar_spec.bar_points("21", 0.0, 3000.0, 500.0, 100.0, "Y",
                                       600.0)
        self.assertEqual(points[1], (500.0, 0.0, 100.0))
        self.assertEqual(points[2], (500.0, 3000.0, 100.0))

    def test_a_leg_with_nowhere_to_go_is_not_bent(self):
        # A nub is not a bend. Below the minimum the bar is placed straight and
        # the layer says why.
        self.assertEqual(len(self.points("21", leg=10.0)), 2)

    def test_leg_height_stops_under_the_top_cover(self):
        # 900 thick, 50 top cover, 16 mm bar: the leg reaches 842, and a bar
        # whose centreline sits at 83 turns up 759.
        self.assertAlmostEqual(
            rebar_spec.leg_height_mm(83.0, 900.0, 50.0, 16.0), 759.0)

    def test_a_top_layer_has_no_room_to_turn_up(self):
        self.assertAlmostEqual(
            rebar_spec.leg_height_mm(842.0, 900.0, 50.0, 16.0), 0.0)

    def test_the_scheduled_shape_reaches_the_bar(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        plans = rebar_spec.plan_footing(data.footing_type("F1"),
                                        data.footing_rebar_for("F1"))
        by_layer = dict((p.row.layer, p) for p in plans)
        self.assertEqual(len(by_layer["B1"].bars[0].points), 4)   # shape 21
        self.assertEqual(len(by_layer["T1"].bars[0].points), 2)   # shape 00

    def test_a_layer_that_cannot_bend_says_so(self):
        footing = models.FootingType("F1", length_mm=3000.0, width_mm=3000.0,
                                     thickness_mm=200.0, cover_top_mm=50.0,
                                     cover_bottom_mm=75.0, cover_side_mm=50.0)
        row = models.FootingRebarRow("F1", layer="B1", direction="X",
                                     diameter_mm=16.0, spacing_mm=200.0,
                                     shape_code="21")
        plan = rebar_spec.plan_footing(footing, [row])[0]
        self.assertEqual(len(plan.bars[0].points), 2)
        self.assertTrue(any("placed straight" in n for n in plan.notes),
                        plan.notes)


class DistributionRegionTests(unittest.TestCase):
    """A set cannot follow a change of slope, so a layer is cut where it turns.

    Told to vary from one end of a tapered pad to the other, Revit interpolates
    straight through the corner and fans bars out past the concrete. One set per
    stretch between the outline's vertices is what a detailer draws and what
    Revit can actually follow.
    """

    #: A house: a rectangle with a gable on top. Apex at x = 2250, eaves at
    #: y = 3000.
    HOUSE = [(0.0, 0.0), (4500.0, 0.0), (4500.0, 3000.0), (2250.0, 4200.0),
             (0.0, 3000.0)]

    def test_breaks_are_the_outline_vertices_on_the_array_axis(self):
        self.assertEqual(rebar_spec.region_breaks(self.HOUSE, "X"),
                         [0.0, 3000.0, 4200.0])
        self.assertEqual(rebar_spec.region_breaks(self.HOUSE, "Y"),
                         [0.0, 2250.0, 4500.0])

    def test_a_rectangle_has_one_region(self):
        square = rebar_spec.rectangle(3000.0, 3000.0)
        positions = [-1400.0, 0.0, 1400.0]
        regions = rebar_spec.split_into_regions(
            positions, rebar_spec.region_breaks(square, "X"))
        self.assertEqual(len(regions), 1)

    def test_positions_are_split_at_an_inner_break(self):
        regions = rebar_spec.split_into_regions(
            [0.0, 1000.0, 2000.0, 3000.0, 4000.0], [0.0, 2500.0, 4000.0])
        self.assertEqual([positions for _label, positions in regions],
                         [[0.0, 1000.0, 2000.0], [3000.0, 4000.0]])

    def test_x_bars_give_a_plain_set_then_a_varying_one(self):
        """Image 5: one normal set across the rectangle, one varying set over
        the gable."""
        row = models.FootingRebarRow("F3", layer="B1", direction="X",
                                     diameter_mm=16.0, spacing_mm=200.0,
                                     shape_code="00")
        plans = rebar_spec.plan_footing_layer(row, self.HOUSE, 900.0, 50.0,
                                              75.0, 50.0)
        self.assertEqual(len(plans), 2)
        self.assertFalse(plans[0].varying)
        self.assertTrue(plans[1].varying)

    def test_y_bars_give_two_varying_sets_either_side_of_the_apex(self):
        """Image 4: the gable rises to the left of the apex and falls to the
        right, so neither side can be one set with the other."""
        row = models.FootingRebarRow("F3", layer="B2", direction="Y",
                                     diameter_mm=16.0, spacing_mm=200.0,
                                     shape_code="00")
        plans = rebar_spec.plan_footing_layer(row, self.HOUSE, 900.0, 50.0,
                                              75.0, 50.0)
        self.assertEqual(len(plans), 2)
        self.assertTrue(all(plan.varying for plan in plans))

    def test_every_bar_survives_the_split(self):
        # Cutting a layer must not lose or duplicate steel.
        row = models.FootingRebarRow("F3", layer="B1", direction="X",
                                     diameter_mm=16.0, spacing_mm=200.0,
                                     shape_code="00")
        plans = rebar_spec.plan_footing_layer(row, self.HOUSE, 900.0, 50.0,
                                              75.0, 50.0)
        positions = []
        for plan in plans:
            positions.extend(bar.points[0][1] for bar in plan.bars)
        self.assertEqual(len(positions), len(set(positions)))
        self.assertEqual(sorted(positions), positions)

    def test_each_region_is_its_own_set(self):
        row = models.FootingRebarRow("F3", layer="B1", direction="X",
                                     diameter_mm=16.0, spacing_mm=200.0,
                                     shape_code="00")
        plans = rebar_spec.plan_footing_layer(row, self.HOUSE, 900.0, 50.0,
                                              75.0, 50.0)
        self.assertTrue(all(plan.element_count == 1 for plan in plans))
        self.assertTrue(all(plan.region for plan in plans))

    def test_a_rectangular_pad_is_not_cut_up(self):
        row = models.FootingRebarRow("F1", layer="B1", direction="X",
                                     diameter_mm=16.0, spacing_mm=200.0,
                                     shape_code="00")
        plans = rebar_spec.plan_footing_layer(
            row, rebar_spec.rectangle(3000.0, 3000.0), 900.0, 50.0, 75.0, 50.0)
        self.assertEqual(len(plans), 1)
        self.assertFalse(plans[0].varying)


class ColumnArrangementTests(unittest.TestCase):

    def arrange(self, count, width=300.0, depth=600.0):
        return rebar_spec.arrange_column_bars(count, width, depth, 40.0, 10.0,
                                              20.0)

    def test_four_bars_are_the_four_corners(self):
        positions = self.arrange(4)
        self.assertEqual(len(positions), 4)
        # 300/2 - (40 + 10 + 10) = 90; 600/2 - 60 = 240
        self.assertEqual(sorted(set(abs(x) for x, _ in positions)), [90.0])
        self.assertEqual(sorted(set(abs(y) for _, y in positions)), [240.0])

    def test_fewer_than_four_still_places_corners(self):
        self.assertEqual(len(self.arrange(2)), 2)

    def test_the_requested_count_is_always_produced(self):
        for count in range(1, 25):
            self.assertEqual(len(self.arrange(count)), count, count)

    def test_extras_favour_the_long_faces(self):
        # A 300 x 600 column should put more intermediate bars down its 600
        # faces than across its 300 ones.
        positions = self.arrange(10)
        on_long_face = [p for p in positions
                        if abs(abs(p[0]) - 90.0) < 1e-6 and abs(p[1]) < 240.0]
        on_short_face = [p for p in positions
                         if abs(abs(p[1]) - 240.0) < 1e-6 and abs(p[0]) < 90.0]
        self.assertGreater(len(on_long_face), len(on_short_face))

    def test_bars_stay_inside_the_section(self):
        for x, y in self.arrange(12):
            self.assertLessEqual(abs(x), 90.0 + 1e-6)
            self.assertLessEqual(abs(y), 240.0 + 1e-6)

    def test_a_section_too_small_for_a_cage_places_nothing(self):
        self.assertEqual(
            rebar_spec.arrange_column_bars(8, 100.0, 100.0, 40.0, 10.0, 20.0),
            [])


class ColumnTieTests(unittest.TestCase):

    def rows(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        return data

    def test_confinement_produces_three_sets(self):
        data = self.rows()
        tie = [r for r in data.column_rebar_for("C1") if r.is_tie][0]
        sets = rebar_spec.plan_column_ties(tie, data.column_type("C1"), 3600.0)
        self.assertEqual(len(sets), 3)
        self.assertEqual([s.spacing_mm for s in sets], [100.0, 200.0, 100.0])

    def test_without_confinement_one_set_runs_the_full_height(self):
        data = self.rows()
        tie = [r for r in data.column_rebar_for("C3") if r.is_tie][0]
        sets = rebar_spec.plan_column_ties(tie, data.column_type("C3"), 3600.0)
        self.assertEqual(len(sets), 1)
        self.assertAlmostEqual(sets[0].array_length_mm, 3600.0)

    def test_a_tie_is_a_closed_loop(self):
        data = self.rows()
        tie = [r for r in data.column_rebar_for("C2") if r.is_tie][0]
        bar = rebar_spec.plan_column_ties(
            tie, data.column_type("C2"), 3000.0)[0].bar
        self.assertEqual(bar.points[0], bar.points[-1])
        self.assertEqual(len(bar.points), 5)

    def test_a_tie_sits_inside_cover_by_half_its_own_diameter(self):
        data = self.rows()
        tie = [r for r in data.column_rebar_for("C2") if r.is_tie][0]
        bar = rebar_spec.plan_column_ties(
            tie, data.column_type("C2"), 3000.0)[0].bar
        # C2 is 400 square, 40 cover, R10 tie: 200 - 40 - 5 = 155
        self.assertAlmostEqual(abs(bar.points[0][0]), 155.0)

    def test_confinement_never_eats_the_whole_column(self):
        data = self.rows()
        tie = [r for r in data.column_rebar_for("C1") if r.is_tie][0]
        sets = rebar_spec.plan_column_ties(tie, data.column_type("C1"), 900.0)
        self.assertTrue(all(s.array_length_mm > 0 for s in sets))


class ColumnMainTests(unittest.TestCase):

    def test_mains_run_the_full_height(self):
        data, _ = excel_engine.parse_grid(fixture_grids())
        row = [r for r in data.column_rebar_for("C2") if r.is_main][0]
        bars = rebar_spec.plan_column_mains(
            row, data.column_type("C2"), 3600.0, tie_diameter_mm=10.0)
        self.assertEqual(len(bars), 8)
        self.assertAlmostEqual(bars[0].length_mm, 3600.0)

    def test_two_main_groups_produce_two_sets_of_bars(self):
        # "4T20 corners + 6T16 faces" is one column and two rows.
        data, _ = excel_engine.parse_grid(fixture_grids())
        mains = [r for r in data.column_rebar_for("C1") if r.is_main]
        counts = [len(rebar_spec.plan_column_mains(
            row, data.column_type("C1"), 3600.0, 10.0)) for row in mains]
        self.assertEqual(sorted(counts), [4, 6])


class EveryFormatTests(unittest.TestCase):
    """The same schedule, in every format, has to become the same objects.

    Regenerate with ``python tests/fixtures/rc_automation/build_fixtures.py``.
    Nothing here needs openpyxl for the text routes; only the workbook ones do.
    """

    def expect(self, data, issues, label):
        self.assertEqual(errors(issues), [], label + ": " + messages(issues))
        self.assertEqual(len(data.footing_types), 3, label)
        self.assertEqual(len(data.footing_placement), 6, label)
        self.assertEqual(len(data.column_rebar), 7, label)
        self.assertEqual(data.footing_type("F1").length_mm, 3000.0, label)
        self.assertEqual(data.units, "mm", label)
        shaped = [p for p in data.footing_placement if p.has_outline]
        self.assertEqual(len(shaped), 1, label)
        self.assertEqual(len(shaped[0].outline), 5, label)

    def test_the_csv_folder(self):
        data, issues = excel_engine.load(_FIXTURES)
        self.expect(data, issues, "csv")

    def test_the_tab_separated_folder(self):
        data, issues = excel_engine.load(os.path.join(_FIXTURES, "txt_sheets"))
        self.expect(data, issues, "txt")
        # The txt set carries the cover sheet, so nothing is assumed.
        self.assertEqual(warnings(issues), [], messages(issues))
        self.assertEqual(data.metadata.get("project"), "Riverside Tower")

    @unittest.skipIf(_openpyxl_missing(), "openpyxl not importable")
    def test_the_xlsx_workbook(self):
        data, issues = excel_engine.load(
            os.path.join(_FIXTURES, "sample_schedule.xlsx"))
        self.expect(data, issues, "xlsx")
        self.assertEqual(warnings(issues), [], messages(issues))

    @unittest.skipIf(_openpyxl_missing(), "openpyxl not importable")
    def test_the_macro_enabled_workbook(self):
        data, issues = excel_engine.load(
            os.path.join(_FIXTURES, "sample_schedule.xlsm"))
        self.expect(data, issues, "xlsm")

    @unittest.skipIf(_openpyxl_missing(), "openpyxl not importable")
    def test_the_workbook_excel_itself_saved(self):
        # Round-tripped through Excel: same schedule, Excel's own bytes.
        data, issues = excel_engine.load(_EXCEL_XLSM)
        self.expect(data, issues, "R1.xlsm")

    def test_the_legacy_workbook_is_refused_with_the_fix(self):
        # A genuine BIFF file saved by Excel, so this is the real refusal path
        # against the real format. The file is present, which makes this the
        # "cannot read this format" message rather than "not found" -- different
        # problems needing different sentences.
        _, issues = excel_engine.load(_LEGACY_XLS)
        self.assertTrue(errors(issues))
        self.assertIn("Save As", errors(issues)[0].message)

    def test_the_legacy_fixture_really_is_the_old_format(self):
        """A .xls placeholder proves nothing; an OLE2 container proves it.

        The first attempt at this fixture was a text file, and it even claimed
        in its own words that Excel would refuse it -- which is false. Excel's
        text import opens a text file named .xls and lays the words out in
        cells. This one starts with the OLE2 compound-document signature, which
        is what a real BIFF workbook is.
        """
        with io.open(_LEGACY_XLS, "rb") as handle:
            signature = handle.read(8)
        self.assertEqual(signature, b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")

    @unittest.skipIf(_openpyxl_missing(), "openpyxl not importable")
    def test_the_workbook_pushed_from_revit_reads(self):
        """The file that produced the first real run, warning and all.

        It has no INFO sheet and no title block, which is exactly why it warned
        about units. Kept as a fixture because a real workbook that exercises
        the fallback is worth more than one written to pass.
        """
        path = os.path.join(
            _FIXTURES, "all-in-one xlsx sheet needed like this example.xlsx")
        data, issues = excel_engine.load(path)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(len(data.footing_types), 3)
        self.assertEqual(len(data.footing_placement), 6)
        found = [i for i in warnings(issues) if "UNITS" in i.message]
        self.assertEqual(len(found), 1, messages(issues))

    @unittest.skipIf(_openpyxl_missing(), "openpyxl not importable")
    def test_text_and_numeric_cells_agree(self):
        """A dimension typed as text and as a number must read the same.

        The pushed workbook stores FOOTING_TYPES as strings and FOOTING_REBAR as
        numbers -- both are what Excel hands back depending on how a cell was
        formatted, and neither may change what gets built.
        """
        pushed, _ = excel_engine.load(os.path.join(
            _FIXTURES, "all-in-one xlsx sheet needed like this example.xlsx"))
        generated, _ = excel_engine.load(
            os.path.join(_FIXTURES, "sample_schedule.xlsx"))
        for mark in ("F1", "F2", "F3"):
            self.assertEqual(pushed.footing_type(mark).length_mm,
                             generated.footing_type(mark).length_mm, mark)
            self.assertEqual(pushed.footing_type(mark).thickness_mm,
                             generated.footing_type(mark).thickness_mm, mark)


class WorkbookFormatTests(unittest.TestCase):
    """The file has to be the format its name claims. Excel checks; readers do not.

    This exists because of a bug these tests originally missed. The generator
    called ``Workbook().save("sample_schedule.xlsm")``, which writes an ordinary
    xlsx and puts an xlsm name on it. openpyxl read it back perfectly -- it goes
    by content and ignores the extension -- so every test passed, while Excel
    compared the declared content type against the extension and refused to open
    the file at all.

    Reading it back is therefore not the check. The check is what the file says
    it is.
    """

    CONTENT_TYPES = {
        ".xlsx": "application/vnd.openxmlformats-officedocument"
                 ".spreadsheetml.sheet.main+xml",
        ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    }

    def declared_type(self, path):
        archive = zipfile.ZipFile(path)
        try:
            text = archive.read("[Content_Types].xml").decode("utf-8")
        finally:
            archive.close()
        marker = 'PartName="/xl/workbook.xml"'
        index = text.find(marker)
        self.assertNotEqual(index, -1, "no workbook part in " + path)
        tail = text[index:]
        key = 'ContentType="'
        start = tail.find(key) + len(key)
        return tail[start:tail.find('"', start)]

    def workbooks(self):
        for name in os.listdir(_FIXTURES):
            extension = os.path.splitext(name)[1].lower()
            if extension in self.CONTENT_TYPES:
                yield os.path.join(_FIXTURES, name), extension

    def test_every_workbook_declares_the_type_its_extension_promises(self):
        checked = 0
        for path, extension in self.workbooks():
            self.assertEqual(self.declared_type(path),
                             self.CONTENT_TYPES[extension],
                             os.path.basename(path))
            checked += 1
        self.assertGreaterEqual(checked, 2, "no workbooks were checked")

    def test_every_workbook_is_a_sound_archive(self):
        for path, _extension in self.workbooks():
            archive = zipfile.ZipFile(path)
            try:
                self.assertIsNone(archive.testzip(), os.path.basename(path))
                self.assertIn("xl/workbook.xml", archive.namelist())
            finally:
                archive.close()

    def test_the_macro_enabled_file_is_not_just_the_xlsx_renamed(self):
        # Identical bytes under two names is exactly the bug, and it is the one
        # thing a content check could still miss if both were wrong together.
        with io.open(os.path.join(_FIXTURES, "sample_schedule.xlsx"), "rb") as h:
            plain = h.read()
        with io.open(os.path.join(_FIXTURES, "sample_schedule.xlsm"), "rb") as h:
            macro = h.read()
        self.assertNotEqual(plain, macro)

    def test_no_placeholder_pretends_to_be_a_workbook(self):
        """Retired placeholders stay retired.

        Both said in their own text that Excel would refuse them. Excel does
        not -- it opens a text file named .xls through its text import. A
        fixture that asserts something untrue about the tool it tests is worse
        than no fixture, and the real Excel-saved .xls replaced them.
        """
        for name in ("not_a_workbook.xls", "sample_schedule.xls"):
            self.assertFalse(os.path.isfile(os.path.join(_FIXTURES, name)),
                             name + " should have been superseded")

    def test_the_generated_xlsm_matches_what_excel_itself_writes(self):
        """Ground truth: Excel's own save of this workbook is in the fixtures.

        openpyxl cannot write an xlsm, so the generator re-declares the workbook
        part by hand. The only way to know that guess is right is to compare it
        with a file Excel actually produced -- and Excel writes exactly the same
        content type, with no vbaProject.bin, for a macro-enabled workbook that
        has no macros in it.
        """
        mine = self.declared_type(os.path.join(_FIXTURES,
                                               "sample_schedule.xlsm"))
        excels = self.declared_type(_EXCEL_XLSM)
        self.assertEqual(mine, excels)
        self.assertEqual(mine, self.CONTENT_TYPES[".xlsm"])

    def test_neither_macro_enabled_file_needs_a_vba_project(self):
        for path in (os.path.join(_FIXTURES, "sample_schedule.xlsm"),
                     _EXCEL_XLSM):
            archive = zipfile.ZipFile(path)
            try:
                self.assertNotIn("xl/vbaProject.bin", archive.namelist(),
                                 os.path.basename(path))
            finally:
                archive.close()


class SingleFileWorkbookTests(unittest.TestCase):
    """Every sheet in one file: one attachment to send, one thing to diff."""

    def test_sheets_are_split_on_the_marker(self):
        rows = excel_engine.split_delimited(
            "#SHEET,FOOTING_TYPES\n"
            "TypeMark,Length\n"
            "F1,3000\n"
            "#SHEET,COLUMN_TYPES\n"
            "TypeMark,Width\n"
            "C1,300\n")
        sheets = excel_engine.split_sheets(rows)
        self.assertEqual(sorted(sheets), ["COLUMN_TYPES", "FOOTING_TYPES"])
        self.assertEqual(sheets["FOOTING_TYPES"][1], ["F1", "3000"])

    def test_rows_before_the_first_marker_are_ignored(self):
        # A file opening with a title should not lose its first sheet to it.
        rows = excel_engine.split_delimited(
            "Riverside Tower schedule,\n"
            "#SHEET,FOOTING_TYPES\n"
            "TypeMark,Length\n"
            "F1,3000\n")
        sheets = excel_engine.split_sheets(rows)
        self.assertEqual(list(sheets), ["FOOTING_TYPES"])
        self.assertEqual(len(sheets["FOOTING_TYPES"]), 2)

    def test_the_marker_is_case_insensitive(self):
        rows = excel_engine.split_delimited("#sheet,INFO\nUNITS,mm\n")
        self.assertEqual(list(excel_engine.split_sheets(rows)), ["INFO"])

    def test_a_marker_with_no_name_starts_nothing(self):
        rows = excel_engine.split_delimited("#SHEET,\nTypeMark,Length\n")
        self.assertEqual(excel_engine.split_sheets(rows), {})

    def test_the_single_csv_reads_like_the_folder(self):
        one_file, one_issues = excel_engine.load(
            os.path.join(_FIXTURES, "sample_schedule.csv"))
        folder, folder_issues = excel_engine.load(_FIXTURES)
        self.assertEqual(errors(one_issues), [], messages(one_issues))
        self.assertEqual(len(one_file.footing_types),
                         len(folder.footing_types))
        self.assertEqual(len(one_file.footing_placement),
                         len(folder.footing_placement))
        self.assertEqual(one_file.footing_type("F1").length_mm,
                         folder.footing_type("F1").length_mm)

    def test_the_single_txt_reads_like_the_folder(self):
        data, issues = excel_engine.load(
            os.path.join(_FIXTURES, "sample_schedule.txt"))
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(warnings(issues), [], messages(issues))
        self.assertEqual(len(data.footing_types), 3)
        self.assertEqual(data.units, "mm")

    def test_the_outline_survives_a_single_file(self):
        # The one cell full of commas, through the marker split as well.
        for name in ("sample_schedule.csv", "sample_schedule.txt"):
            data, _ = excel_engine.load(os.path.join(_FIXTURES, name))
            shaped = [p for p in data.footing_placement if p.has_outline]
            self.assertEqual(len(shaped), 1, name)
            self.assertEqual(len(shaped[0].outline), 5, name)


class DelimitedTextTests(unittest.TestCase):
    """Delimited text is a real input: Revit's own schedule export writes it."""

    def test_a_comma_file_splits(self):
        rows = excel_engine.split_delimited("a,b,c\n1,2,3")
        self.assertEqual(rows, [["a", "b", "c"], ["1", "2", "3"]])

    def test_a_tab_header_switches_the_delimiter(self):
        rows = excel_engine.split_delimited("a\tb\tc\n1\t2\t3")
        self.assertEqual(rows, [["a", "b", "c"], ["1", "2", "3"]])

    def test_a_quoted_cell_keeps_its_commas(self):
        # The Outline column is one cell full of commas and semicolons;
        # splitting naively would tear a pad's shape into six columns.
        rows = excel_engine.split_delimited(
            'Mark,Outline\nF3,"0,0; 4500,0; 4500,3000"')
        self.assertEqual(rows[1], ["F3", "0,0; 4500,0; 4500,3000"])

    def test_a_doubled_quote_is_one_quote(self):
        rows = excel_engine.split_delimited('a,b\n1,"say ""hi"""')
        self.assertEqual(rows[1], ["1", 'say "hi"'])

    def test_windows_line_endings(self):
        rows = excel_engine.split_delimited("a,b\r\n1,2\r\n")
        self.assertEqual(rows, [["a", "b"], ["1", "2"]])

    def test_a_blank_line_is_a_blank_row_not_a_ragged_one(self):
        rows = excel_engine.split_delimited("a,b\n\n1,2")
        self.assertEqual(rows, [["a", "b"], [], ["1", "2"]])


class TextFolderTests(unittest.TestCase):
    """A folder of sheets reads exactly like the workbook it came from."""

    def test_the_csv_fixtures_load_as_a_workbook(self):
        data, issues = excel_engine.load(_FIXTURES)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(len(data.footing_types), 3)
        self.assertEqual(len(data.footing_placement), 6)
        self.assertEqual(data.footing_type("F1").length_mm, 3000.0)

    def test_the_outline_survives_the_folder_route(self):
        data, _ = excel_engine.load(_FIXTURES)
        shaped = [p for p in data.footing_placement if p.has_outline]
        self.assertEqual(len(shaped), 1)
        self.assertEqual(len(shaped[0].outline), 5)

    def test_a_folder_with_no_sheets_names_what_it_wanted(self):
        empty = tempfile.mkdtemp(prefix="rc_empty_")
        try:
            _, issues = excel_engine.load(empty)
            message = errors(issues)[0].message
            self.assertIn("No sheets found", message)
            self.assertIn("FOOTING_TYPES", message)
        finally:
            shutil.rmtree(empty, ignore_errors=True)

    def test_a_stray_file_beside_the_sheets_is_ignored_not_fatal(self):
        """A folder is somebody's working directory as often as a workbook.

        An exported report sitting beside the sheets used to be read as one and
        killed the whole load — it is not UTF-8, because Excel on Windows does
        not write UTF-8.
        """
        data, issues = excel_engine.load(_FIXTURES)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(len(data.footing_types), 3)
        notes = [i for i in issues if i.severity == models.SEVERITY_INFO]
        self.assertTrue(any("not named after a sheet" in i.message
                            for i in notes), messages(issues))

    def test_a_cp1252_sheet_reads(self):
        """Excel on a Western Windows writes cp1252, not UTF-8."""
        folder = tempfile.mkdtemp(prefix="rc_cp1252_")
        try:
            path = os.path.join(folder, "FOOTING_TYPES.csv")
            with open(path, "wb") as handle:
                handle.write(u"TypeMark,Length,Width,Thickness,CoverTop,"
                             u"CoverBottom,CoverSide,Comments\n"
                             u"F1,3000,3000,900,50,75,50,pad \u2014 typical\n"
                             .encode("cp1252"))
            grids, issues = excel_engine.read_grid(folder)
            self.assertEqual(errors(issues), [], messages(issues))
            self.assertIn(u"\u2014", grids["FOOTING_TYPES"][1][7])
        finally:
            shutil.rmtree(folder, ignore_errors=True)


class CoverSheetTests(unittest.TestCase):
    """Metadata belongs on its own sheet, so the data sheets stay tables."""

    def info(self, rows):
        grids = minimal_grids()
        grids["INFO"] = grid(rows)
        return grids

    def test_the_cover_sheet_declares_the_units(self):
        data, issues = excel_engine.parse_grid(self.info("""
PROJECT,Riverside Tower
UNITS,mm
STANDARD,BS 8666:2020
"""))
        self.assertEqual(data.units, "mm")
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(warnings(issues), [], messages(issues))
        self.assertEqual(data.metadata.get("project"), "Riverside Tower")

    def test_the_sheet_may_be_called_something_else(self):
        for name in ("Project Info", "COVER", "Settings"):
            grids = minimal_grids()
            grids[name] = grid("UNITS,mm")
            data, issues = excel_engine.parse_grid(grids)
            self.assertEqual(warnings(issues), [], name + ": " + messages(issues))
            self.assertEqual(data.units, "mm", name)

    def test_wrong_units_on_the_cover_sheet_are_still_refused(self):
        _, issues = excel_engine.parse_grid(self.info("UNITS,m"))
        self.assertTrue(any("UNITS" in i.message for i in errors(issues)),
                        messages(issues))

    def test_blank_and_label_only_rows_are_skipped(self):
        data, _ = excel_engine.parse_grid(self.info("""
Project schedule,
,
UNITS,mm
"""))
        self.assertEqual(data.units, "mm")

    def test_the_warning_names_the_sheet_that_fixes_it(self):
        # This is the workbook that produced the only warning on the first real
        # run in Revit: pure data sheets, no title block, nothing declaring mm.
        _, issues = excel_engine.parse_grid(minimal_grids(FOOTING_TYPES=grid("""
TypeMark,Length,Width,Thickness,CoverTop,CoverBottom,CoverSide
F1,3000,3000,900,50,75,50
""")))
        found = [i for i in warnings(issues) if "UNITS" in i.message]
        self.assertTrue(found, messages(issues))
        self.assertIn("INFO", found[0].message)

    def test_a_title_block_still_works_without_a_cover_sheet(self):
        # A real schedule arrives with its title block above the header, and
        # rejecting that would make the tool useless on the documents it exists
        # to read.
        data, issues = excel_engine.parse_grid(fixture_grids())
        self.assertEqual(warnings(issues), [], messages(issues))
        self.assertEqual(data.metadata.get("project"), "Riverside Tower")


# ── the file layer ─────────────────────────────────────────────────────────

try:
    import openpyxl as _openpyxl
except Exception:                       # pragma: no cover - platform dependent
    _openpyxl = None


@unittest.skipIf(_openpyxl is None,
                 "openpyxl not importable — the extension vendors a Windows build")
class ReadGridTests(unittest.TestCase):
    """``read_grid`` is thin, but it is the only thing standing between a real
    workbook and everything else, so its failure modes are worth pinning."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix="rc_automation_")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def write_workbook(self, grids, name="schedule.xlsx"):
        book = _openpyxl.Workbook()
        book.remove(book.active)
        for sheet_name, rows in grids.items():
            sheet = book.create_sheet(title=sheet_name)
            for row in rows:
                sheet.append(list(row))
        path = os.path.join(self.tmp, name)
        book.save(path)
        return path

    def test_a_real_workbook_round_trips_to_the_same_objects(self):
        # The CSV fixtures and a genuine .xlsx must parse identically, which is
        # what makes testing everything else off a CSV legitimate.
        path = self.write_workbook(fixture_grids())
        data, issues = excel_engine.load(path)
        self.assertEqual(errors(issues), [], messages(issues))
        self.assertEqual(len(data.footing_types), 3)
        self.assertEqual(data.footing_type("F1").length_mm, 3000.0)
        self.assertEqual(data.metadata.get("project"), "Riverside Tower")

    def test_native_excel_types_read_the_same_as_text(self):
        # openpyxl hands back ints and floats where the CSV hands back strings;
        # a shape code of 0 and a count of 15.0 must survive both routes.
        path = self.write_workbook({
            "FOOTING_TYPES": [["UNITS", "mm"],
                              ["TypeMark", "Length", "Width", "Thickness",
                               "CoverTop", "CoverBottom", "CoverSide"],
                              ["F1", 3000, 3000, 900, 50, 75, 50]],
            "FOOTING_REBAR": [["TypeMark", "Layer", "Direction", "Diameter",
                               "Count", "Spacing", "ShapeCode"],
                              ["F1", "B1", "X", 16, 15.0, 200, 0]],
            "COLUMN_TYPES": [["TypeMark", "Width", "Depth", "Cover"],
                             ["C1", 400, 400, 40]],
            "COLUMN_REBAR": [["TypeMark", "BarRole", "Diameter", "Count",
                              "Spacing", "ShapeCode"],
                             ["C1", "Main", 20, 8, None, 0],
                             ["C1", "Tie", 10, None, 200, 51]],
            "FOOTING_PLACEMENT": [["Mark", "TypeMark", "X", "Y", "Level"],
                                  ["F1-A1", "F1", 12500, 8400.5, "Foundation"]],
            "COLUMN_PLACEMENT": [["Mark", "TypeMark", "X", "Y", "BaseLevel",
                                  "TopLevel"],
                                 ["C1-A1", "C1", 12500, 8400.5, "Foundation",
                                  "Level 1"]],
        })
        data, issues = excel_engine.load(path)
        self.assertEqual(errors(issues), [], messages(issues))
        row = data.footing_rebar[0]
        self.assertEqual(row.count, 15)
        self.assertEqual(row.shape_code, "00")
        self.assertEqual(data.column_rebar[1].shape_code, "51")
        # Native numeric coordinates survive the real Excel path.
        self.assertEqual(data.footing_placement[0].x_mm, 12500.0)
        self.assertEqual(data.footing_placement[0].y_mm, 8400.5)

    def test_a_missing_file_is_a_sentence_not_a_traceback(self):
        data, issues = excel_engine.load(os.path.join(self.tmp, "nope.xlsx"))
        self.assertTrue(errors(issues))
        self.assertIn("not found", errors(issues)[0].message)
        self.assertTrue(data.is_empty())

    def test_legacy_xls_says_what_to_do_about_it(self):
        path = os.path.join(self.tmp, "old.xls")
        with io.open(path, "w") as handle:
            handle.write(u"not really a workbook")
        _, issues = excel_engine.load(path)
        self.assertIn("Save As", errors(issues)[0].message)

    def test_a_file_that_is_not_a_workbook_is_refused_by_extension(self):
        _, issues = excel_engine.load(os.path.join(self.tmp, "notes.docx"))
        self.assertTrue(errors(issues))
        self.assertIn("Excel workbook", errors(issues)[0].message)

    def test_a_text_file_with_no_sheet_markers_says_both_ways_out(self):
        # One unmarked table is one sheet, and a schedule needs six. The message
        # has to name both layouts that work, not just refuse.
        path = os.path.join(self.tmp, "FOOTING_TYPES.txt")
        with io.open(path, "w") as handle:
            handle.write(u"TypeMark\tLength\nF1\t3000\n")
        _, issues = excel_engine.load(path)
        message = errors(issues)[0].message
        self.assertIn("#SHEET", message)
        self.assertIn("folder", message)

    def test_no_path_is_refused(self):
        _, issues = excel_engine.load("")
        self.assertTrue(errors(issues))

    def test_a_corrupt_workbook_is_reported_not_raised(self):
        path = os.path.join(self.tmp, "corrupt.xlsx")
        with io.open(path, "w") as handle:
            handle.write(u"PK\x03\x04 and then nonsense")
        _, issues = excel_engine.load(path)
        self.assertTrue(errors(issues), "a corrupt file must not raise")

    def test_reading_does_not_hold_the_file_open(self):
        # read_only workbooks lock the file until closed, which would lock the
        # user out of their own schedule in Excel.
        path = self.write_workbook(fixture_grids())
        excel_engine.load(path)
        os.remove(path)
        self.assertFalse(os.path.exists(path))


if __name__ == "__main__":
    unittest.main()
