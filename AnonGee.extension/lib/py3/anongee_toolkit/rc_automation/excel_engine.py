# -*- coding: utf-8 -*-
"""Read an RC Automation workbook, and turn its cells into objects.

The module is deliberately in two halves, and the seam matters.

:func:`read_grid` is the only function that knows openpyxl exists. It opens the
file, pulls every sheet out as a list of lists of raw Python values, and stops.
It is small and dull because it cannot be unit-tested anywhere openpyxl will not
install -- the extension vendors a Windows build of numpy, so the machine that
runs the test suite generally cannot import it at all.

:func:`parse_grid` does everything else -- finding the header row, matching
column names, coercing cells, building the objects -- from those plain lists.
It has no dependencies whatsoever, so every rule it applies is testable with a
hand-written list of rows on any machine. That is where the behaviour lives, and
that is where the tests point.

Two things this module will not do:

*It will not raise at the user.* A workbook that cannot be read produces an
``Issue`` with a sentence the user can act on, because a traceback in a modeless
window is a support call.

*It will not use* ``re``. The CPython 3 engine ships a partial stdlib and ``re``
is one of the missing modules (§12.9.3). Header matching is done by folding case
and dropping non-alphanumerics, which is all it ever needed.
"""

from anongee_toolkit.rc_automation import models
from anongee_toolkit.rc_automation import standards
from anongee_toolkit.rc_automation.models import (
    Issue, SEVERITY_ERROR, SEVERITY_WARNING, SEVERITY_INFO)

__version__ = "0.1.0"

#: Rows to search for a header before giving up. Real schedules carry a title
#: block -- project, drawing number, revision -- above the table, and the header
#: is found rather than assumed so those files load without being edited first.
_MAX_HEADER_SCAN_ROWS = 12

#: openpyxl reads neither of these. ``.xls`` is the pre-2007 binary format and
#: the message has to say so, because "could not read the file" sends the user
#: looking for corruption instead of Save As.
_LEGACY_EXTENSIONS = (".xls", ".xlsb")
_READABLE_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")

#: Delimited text. Revit's own schedule export writes tab-separated text, so
#: this is a first-class input rather than a testing convenience.
#:
#: Two layouts, because both are how people actually keep these. A *folder* of
#: files is one sheet per file, named by the file. A *single file* carries every
#: sheet, separated by :data:`SHEET_MARKER` rows -- one attachment to send,
#: one thing to diff, and it still opens in Excel as a readable column.
_TEXT_EXTENSIONS = (".csv", ".tsv", ".txt")

#: The first cell of a row that starts a new sheet in a single-file workbook.
#: The second cell is the sheet name: ``#SHEET,FOOTING_TYPES``.
#:
#: A leading ``#`` because no column heading or type mark begins with one, so it
#: cannot collide with data, and Excel shows it as an ordinary row rather than
#: choking on it.
SHEET_MARKER = "#SHEET"


# ---------------------------------------------------------------------------
# Text folding
# ---------------------------------------------------------------------------

def fold(text):
    """Lowercase, alphanumerics only -- the key both sheets and columns match on.

    ``"Cover Top"``, ``"cover_top"`` and ``"CoverTop"`` all fold to
    ``"covertop"``, so a workbook is not rejected over a space.
    """
    if text is None:
        return ""
    return "".join(ch.lower() for ch in str(text) if ch.isalnum())


def coerce_text(value):
    """A cell as trimmed text. ``None`` and blanks both become ``""``."""
    if value is None:
        return ""
    return str(value).strip()


def coerce_number(value):
    """``(number, error)`` -- a float, or a sentence saying why it is not one.

    Blank is not an error; it is ``(None, None)``, and whether blank is allowed
    is the caller's rule, not this function's. Content that is not a number is
    an error naming what was found, since "expected a number, found 'N/A'" is
    the message that gets the cell fixed.

    Thousands separators and stray unit suffixes are the two things real
    schedules actually contain, so ``"1,200"`` and ``"1200 mm"`` both read.
    """
    if value is None:
        return None, None
    if isinstance(value, bool):
        return None, "expected a number, found {0}".format(value)
    if isinstance(value, (int, float)):
        return float(value), None

    text = str(value).strip()
    if not text:
        return None, None

    cleaned = text.replace(",", "").replace(" ", "")
    for suffix in ("mm", "MM", "Mm"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[:-len(suffix)]
            break
    if not cleaned:
        return None, "expected a number, found {0!r}".format(text)
    try:
        return float(cleaned), None
    except ValueError:
        return None, "expected a number, found {0!r}".format(text)


def coerce_count(value):
    """``(int, error)`` for a bar count -- whole and positive, or nothing.

    ``12.0`` is accepted because that is how Excel hands back a cell holding 12.
    ``12.5`` is not, because half a bar is a typo and rounding it hides one.
    """
    number, error = coerce_number(value)
    if error or number is None:
        return None, error
    if abs(number - round(number)) > 1e-9:
        return None, "expected a whole number of bars, found {0}".format(number)
    return int(round(number)), None


def parse_outline(value):
    """``(points, error)`` from ``"0,0; 3000,0; 3000,2000; 0,2000"``.

    An arbitrary pad outline, millimetres relative to the placement point and
    before rotation. This is the column that earns footings being floors rather
    than family instances: a combined pad, a cut corner or a footing worked
    around a pile cap can be scheduled instead of approximated by the nearest
    rectangle.

    Blank is not an error -- most footings are rectangles and say so through the
    type's Length and Width.
    """
    text = coerce_text(value)
    if not text:
        return None, None

    points = []
    for chunk in text.replace("\n", ";").split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        parts = chunk.split(",")
        if len(parts) != 2:
            return None, ("outline point {0!r} is not an x,y pair".format(chunk))
        x, x_error = coerce_number(parts[0])
        y, y_error = coerce_number(parts[1])
        if x_error or y_error or x is None or y is None:
            return None, "outline point {0!r} is not a pair of numbers".format(chunk)
        points.append((x, y))

    if len(points) < 3:
        return None, ("an outline needs at least 3 points, found {0}".format(
            len(points)))
    # A closing point repeating the first is how people write these; the curve
    # loop is built by walking back to the start, so carrying it would place a
    # zero-length line and Revit would refuse the whole sketch.
    if len(points) > 3 and points[0] == points[-1]:
        points = points[:-1]
    return points, None


# ---------------------------------------------------------------------------
# Column specifications
# ---------------------------------------------------------------------------
# (attribute, canonical header, required, kind, aliases)
#
# Aliases exist so the names in the original feature specification -- and the
# abbreviations engineers actually type -- load without anyone editing a file.
_TEXT, _NUM, _COUNT, _SHAPE, _OUTLINE = (
    "text", "number", "count", "shape", "outline")

_FOOTING_TYPE_COLUMNS = (
    ("type_mark",      "TypeMark",    True,  _TEXT, ("Mark", "FootingMark", "Type")),
    ("length_mm",      "Length",      True,  _NUM,  ("L",)),
    ("width_mm",       "Width",       True,  _NUM,  ("B", "W")),
    ("thickness_mm",   "Thickness",   True,  _NUM,  ("Depth", "D", "T")),
    ("cover_top_mm",   "CoverTop",    True,  _NUM,  ("TopCover",)),
    ("cover_bottom_mm", "CoverBottom", True, _NUM,  ("BottomCover",)),
    ("cover_side_mm",  "CoverSide",   True,  _NUM,  ("SideCover",)),
    ("concrete",       "Concrete",    False, _TEXT, ("Grade", "ConcreteGrade")),
    ("comments",       "Comments",    False, _TEXT, ("Remarks", "Note", "Notes")),
)

_COLUMN_TYPE_COLUMNS = (
    ("type_mark",  "TypeMark",  True,  _TEXT, ("Mark", "ColumnMark", "Type")),
    ("width_mm",   "Width",     True,  _NUM,  ("B", "W")),
    ("depth_mm",   "Depth",     True,  _NUM,  ("D", "H")),
    ("cover_mm",   "Cover",     True,  _NUM,  ("ClearCover",)),
    ("concrete",   "Concrete",  False, _TEXT, ("Grade", "ConcreteGrade")),
    ("comments",   "Comments",  False, _TEXT, ("Remarks", "Note", "Notes")),
)

_FOOTING_REBAR_COLUMNS = (
    ("type_mark",    "TypeMark",  True,  _TEXT,  ("FootingMark", "Mark")),
    ("layer",        "Layer",     True,  _TEXT,  ("Position",)),
    ("direction",    "Direction", True,  _TEXT,  ("Dir", "Axis")),
    ("bar_type",     "BarType",   False, _TEXT,  ("Type", "RebarType", "Grade")),
    ("diameter_mm",  "Diameter",  True,  _NUM,   ("Dia", "Phi", "Size")),
    ("count",        "Count",     False, _COUNT, ("Nos", "No", "Number", "Qty")),
    ("spacing_mm",   "Spacing",   False, _NUM,   ("Pitch", "CC", "Centres")),
    ("shape_code",   "ShapeCode", True,  _SHAPE, ("Shape", "Code")),
    ("end_cover_mm", "EndCover",  False, _NUM,   ("SideCover",)),
    ("comments",     "Comments",  False, _TEXT,  ("Remarks", "Note", "Notes")),
)

_COLUMN_REBAR_COLUMNS = (
    ("type_mark",             "TypeMark",          True,  _TEXT,  ("ColumnMark", "Mark")),
    ("bar_role",              "BarRole",           True,  _TEXT,  ("Role", "BarKind")),
    ("bar_type",              "BarType",           False, _TEXT,  ("Type", "RebarType", "Grade")),
    ("diameter_mm",           "Diameter",          True,  _NUM,   ("Dia", "Phi", "Size")),
    ("count",                 "Count",             False, _COUNT, ("Nos", "No", "Number", "Qty")),
    ("spacing_mm",            "Spacing",           False, _NUM,   ("Pitch", "CC", "Centres")),
    ("shape_code",            "ShapeCode",         True,  _SHAPE, ("Shape", "Code")),
    ("spacing_end_mm",        "SpacingEnd",        False, _NUM,   ("EndSpacing", "ConfinedSpacing")),
    ("confinement_length_mm", "ConfinementLength", False, _NUM,   ("Lo", "ConfinementZone")),
    ("comments",              "Comments",          False, _TEXT,  ("Remarks", "Note", "Notes")),
)

_FOOTING_PLACEMENT_COLUMNS = (
    ("mark",          "Mark",      True,  _TEXT,    ("FootingMark", "Instance")),
    ("type_mark",     "TypeMark",  True,  _TEXT,    ("Type", "FootingType")),
    ("grid_x",        "GridX",     False, _TEXT,    ("GridRefX", "AxisX")),
    ("grid_y",        "GridY",     False, _TEXT,    ("GridRefY", "AxisY")),
    ("x_mm",          "X",         False, _NUM,     ("Easting", "XCoord")),
    ("y_mm",          "Y",         False, _NUM,     ("Northing", "YCoord")),
    ("level",         "Level",     True,  _TEXT,    ("BaseLevel", "Storey")),
    ("top_offset_mm", "TopOffset", False, _NUM,     ("Offset",)),
    ("rotation_deg",  "Rotation",  False, _NUM,     ("Angle",)),
    ("outline",       "Outline",   False, _OUTLINE, ("Shape", "Polygon")),
)

_COLUMN_PLACEMENT_COLUMNS = (
    ("mark",           "Mark",       True,  _TEXT, ("ColumnMark", "Instance")),
    ("type_mark",      "TypeMark",   True,  _TEXT, ("Type", "ColumnType")),
    ("grid_x",         "GridX",      False, _TEXT, ("GridRefX", "AxisX")),
    ("grid_y",         "GridY",      False, _TEXT, ("GridRefY", "AxisY")),
    ("x_mm",           "X",          False, _NUM,  ("Easting", "XCoord")),
    ("y_mm",           "Y",          False, _NUM,  ("Northing", "YCoord")),
    ("base_level",     "BaseLevel",  True,  _TEXT, ("Level", "FromLevel")),
    ("base_offset_mm", "BaseOffset", False, _NUM,  ("BottomOffset",)),
    ("top_level",      "TopLevel",   True,  _TEXT, ("ToLevel",)),
    ("top_offset_mm",  "TopOffset",  False, _NUM,  ("UpperOffset",)),
    ("rotation_deg",   "Rotation",   False, _NUM,  ("Angle",)),
)

_SHEET_SPECS = (
    (models.SHEET_FOOTING_TYPES, _FOOTING_TYPE_COLUMNS, models.FootingType),
    (models.SHEET_COLUMN_TYPES, _COLUMN_TYPE_COLUMNS, models.ColumnType),
    (models.SHEET_FOOTING_REBAR, _FOOTING_REBAR_COLUMNS, models.FootingRebarRow),
    (models.SHEET_COLUMN_REBAR, _COLUMN_REBAR_COLUMNS, models.ColumnRebarRow),
    (models.SHEET_FOOTING_PLACEMENT, _FOOTING_PLACEMENT_COLUMNS,
     models.FootingPlacement),
    (models.SHEET_COLUMN_PLACEMENT, _COLUMN_PLACEMENT_COLUMNS,
     models.ColumnPlacement),
)


# ---------------------------------------------------------------------------
# Half one: the file
# ---------------------------------------------------------------------------

def split_delimited(text):
    """Rows from delimited text, working out the delimiter from the header.

    Hand-rolled because the engine has no ``csv`` module (§12.9.3), and quote
    aware because the Outline column is a single cell full of commas and
    semicolons -- splitting naively would tear one pad's shape into six columns.

    Tab wins where the header has one, since that is what Revit's own schedule
    export writes; otherwise comma.
    """
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    if not lines:
        return []
    delimiter = "\t" if "\t" in lines[0] else ","

    rows = []
    for line in lines:
        if not line.strip():
            rows.append([])
            continue
        cells = []
        current = []
        quoted = False
        index = 0
        while index < len(line):
            char = line[index]
            if char == '"':
                if quoted and index + 1 < len(line) and line[index + 1] == '"':
                    current.append('"')      # "" inside a quoted cell is one "
                    index += 1
                else:
                    quoted = not quoted
            elif char == delimiter and not quoted:
                cells.append("".join(current))
                current = []
            else:
                current.append(char)
            index += 1
        cells.append("".join(current))
        rows.append(cells)

    while rows and not rows[-1]:
        rows.pop()
    return rows


def split_sheets(rows):
    """``{sheet name: rows}`` from one file holding several sheets.

    Rows before the first marker are ignored: a file that opens with a title or
    a note should not lose its first sheet to it, and a stray line is a likelier
    explanation than an unnamed sheet.
    """
    sheets = {}
    current = None
    for row in rows:
        if row and coerce_text(row[0]).upper() == SHEET_MARKER:
            name = coerce_text(row[1]) if len(row) > 1 else ""
            current = name or None
            if current:
                sheets.setdefault(current, [])
            continue
        if current is not None:
            sheets[current].append(row)
    return sheets


def read_text_file(path):
    """``({sheet: rows}, issues)`` from a single delimited file."""
    issues = []
    try:
        handle = open(path, "r")
        try:
            text = handle.read()
        finally:
            handle.close()
    except Exception as read_error:
        issues.append(Issue(
            SEVERITY_ERROR, "Could not read {0}: {1}".format(path, read_error)))
        return {}, issues

    sheets = split_sheets(split_delimited(text))
    if not sheets:
        issues.append(Issue(
            SEVERITY_ERROR,
            "This file holds one table, and a schedule needs several sheets. "
            "Either separate them with '{0}' rows naming each sheet, or put one "
            "file per sheet in a folder and select the folder.".format(
                SHEET_MARKER)))
    return sheets, issues


def read_text_folder(path):
    """``({sheet name: rows}, issues)`` from a folder of delimited text files."""
    import os
    issues = []
    grids = {}
    try:
        names = sorted(os.listdir(path))
    except Exception as list_error:
        issues.append(Issue(
            SEVERITY_ERROR,
            "Could not read the folder: {0}".format(list_error)))
        return {}, issues

    for name in names:
        stem, extension = os.path.splitext(name)
        if extension.lower() not in _TEXT_EXTENSIONS:
            continue
        try:
            handle = open(os.path.join(path, name), "r")
            try:
                grids[stem] = split_delimited(handle.read())
            finally:
                handle.close()
        except Exception as read_error:
            issues.append(Issue(
                SEVERITY_ERROR,
                "Could not read {0}: {1}".format(name, read_error)))

    if not grids:
        issues.append(Issue(
            SEVERITY_ERROR,
            "No .csv, .tsv or .txt sheets in {0}.".format(path)))
    return grids, issues


def read_grid(path):
    """``({sheet_name: [[cell, ...], ...]}, issues)`` -- every sheet, raw.

    Deliberately the whole of the openpyxl surface. Values come back exactly as
    openpyxl hands them over: ints, floats, strings, datetimes, ``None``.
    Interpretation belongs to :func:`parse_grid`.

    ``data_only=True`` asks for cached formula results rather than formula text.
    Excel writes that cache; other writers frequently do not, so a workbook
    generated by a non-Excel tool can come back as a grid of ``None`` -- which
    :func:`parse_grid` reports as an empty sheet, and which is why the caller
    also gets told when a sheet parsed to nothing.
    """
    issues = []

    if not path:
        issues.append(Issue(SEVERITY_ERROR, "No workbook selected."))
        return {}, issues

    import os
    if os.path.isdir(path):
        return read_text_folder(path)

    lowered = str(path).lower()
    if any(lowered.endswith(extension) for extension in _TEXT_EXTENSIONS):
        if not os.path.isfile(path):
            issues.append(Issue(
                SEVERITY_ERROR, "File not found: {0}".format(path)))
            return {}, issues
        return read_text_file(path)
    for extension in _LEGACY_EXTENSIONS:
        if lowered.endswith(extension):
            issues.append(Issue(
                SEVERITY_ERROR,
                "{0} files cannot be read. Open the file in Excel and use "
                "Save As to write a .xlsx, then load that.".format(extension)))
            return {}, issues
    if not any(lowered.endswith(ext) for ext in _READABLE_EXTENSIONS):
        issues.append(Issue(
            SEVERITY_ERROR,
            "Expected an Excel workbook (.xlsx or .xlsm), or a folder of .csv "
            "/ .tsv / .txt sheets. Got {0!r}.".format(path)))
        return {}, issues

    if not os.path.isfile(path):
        issues.append(Issue(
            SEVERITY_ERROR, "Workbook not found: {0}".format(path)))
        return {}, issues

    try:
        from openpyxl import load_workbook
    except Exception as import_error:
        issues.append(Issue(
            SEVERITY_ERROR,
            "openpyxl is not available, so no workbook can be read ({0}). "
            "Check the extension's lib/py3 folder.".format(import_error)))
        return {}, issues

    workbook = None
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        grids = {}
        for name in workbook.sheetnames:
            sheet = workbook[name]
            grids[name] = [list(row) for row in sheet.iter_rows(values_only=True)]
        return grids, issues
    except Exception as read_error:
        issues.append(Issue(
            SEVERITY_ERROR,
            "Could not read the workbook: {0}".format(read_error)))
        return {}, issues
    finally:
        # read_only workbooks hold the file open until closed, and leaving it
        # open would lock the user out of their own schedule in Excel.
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Half two: the grid
# ---------------------------------------------------------------------------

def _row_is_blank(row):
    return all(coerce_text(cell) == "" for cell in row)


def find_header_row(rows, columns):
    """Index of the first row carrying every required header, or ``None``.

    Searching rather than assuming row 1 is what lets a real schedule load: the
    title block above the table is normal, and rejecting the file over it would
    make the tool useless on the documents it exists to read.
    """
    required = set()
    for _attr, header, is_required, _kind, aliases in columns:
        if is_required:
            required.add(fold(header))
    if not required:
        return None

    limit = min(len(rows), _MAX_HEADER_SCAN_ROWS)
    for index in range(limit):
        folded = set(fold(cell) for cell in rows[index] if coerce_text(cell))
        matched = 0
        for _attr, header, is_required, _kind, aliases in columns:
            if not is_required:
                continue
            names = [fold(header)] + [fold(alias) for alias in aliases]
            if any(name in folded for name in names):
                matched += 1
        if matched == len(required):
            return index
    return None


def read_key_value_sheet(rows):
    """``{folded key: text}`` from a two-column cover sheet.

    Column A is the label, column B the value. Blank rows and label-only rows
    are skipped, so a sheet laid out with spacing and section headings still
    reads.
    """
    metadata = {}
    for row in rows or ():
        if len(row) < 2:
            continue
        key = fold(row[0])
        value = coerce_text(row[1])
        if key and value:
            metadata.setdefault(key, value)
    return metadata


def read_metadata(rows, header_index):
    """``{folded key: text}`` from the rows above the header.

    A title block is two columns wide in practice -- a label and a value -- so
    that is what is read. It is how ``UNITS | mm`` reaches the validator.
    """
    metadata = {}
    for index in range(min(header_index, len(rows))):
        row = rows[index]
        if len(row) < 2:
            continue
        key = fold(row[0])
        value = coerce_text(row[1])
        if key and value:
            metadata[key] = value
    return metadata


def _column_index_map(header_row, columns, sheet_name, issues):
    """``({attribute: column index}, ok)``, reporting each required column missing.

    ``ok`` is False when a required column was not found. The rows below it
    cannot be read, and saying so once beats one identical complaint per row.
    """
    folded = {}
    for index, cell in enumerate(header_row):
        key = fold(cell)
        # First occurrence wins: a workbook with a stray repeated header should
        # read its leftmost real column, not the empty copy someone pasted.
        if key and key not in folded:
            folded[key] = index

    mapping = {}
    complete = True
    for attr, header, is_required, _kind, aliases in columns:
        names = [fold(header)] + [fold(alias) for alias in aliases]
        for name in names:
            if name in folded:
                mapping[attr] = folded[name]
                break
        else:
            if is_required:
                complete = False
                issues.append(Issue(
                    SEVERITY_ERROR,
                    "Required column {0!r} is missing.".format(header),
                    sheet=sheet_name))
    return mapping, complete


def _cell(row, index):
    return row[index] if index is not None and index < len(row) else None


def _read_row(row, mapping, columns, sheet_name, row_number, issues):
    """One data row into a kwargs dict, one Issue per cell that will not read."""
    values = {}
    for attr, header, _required, kind, _aliases in columns:
        raw = _cell(row, mapping.get(attr))
        if kind == _TEXT:
            values[attr] = coerce_text(raw)
        elif kind == _OUTLINE:
            points, error = parse_outline(raw)
            if error:
                issues.append(Issue(
                    SEVERITY_ERROR, error,
                    sheet=sheet_name, row=row_number, column=header))
            values[attr] = points
        elif kind == _SHAPE:
            code = standards.normalise_shape_code(raw)
            if code is None and coerce_text(raw):
                issues.append(Issue(
                    SEVERITY_ERROR,
                    "{0!r} is not a shape code.".format(coerce_text(raw)),
                    sheet=sheet_name, row=row_number, column=header))
            values[attr] = code
        else:
            coerce = coerce_count if kind == _COUNT else coerce_number
            number, error = coerce(raw)
            if error:
                issues.append(Issue(
                    SEVERITY_ERROR, error,
                    sheet=sheet_name, row=row_number, column=header))
            values[attr] = number
    return values


def _parse_sheet(rows, columns, factory, sheet_name, issues):
    """``(objects, metadata)`` for one sheet, or ``([], {})`` if unreadable."""
    if not rows or all(_row_is_blank(row) for row in rows):
        issues.append(Issue(
            SEVERITY_ERROR,
            "Sheet is empty. If the schedule is built from formulas, open it in "
            "Excel and save it once so the values are stored.",
            sheet=sheet_name))
        return [], {}

    header_index = find_header_row(rows, columns)
    if header_index is None:
        wanted = ", ".join(header for _a, header, req, _k, _al in columns if req)
        issues.append(Issue(
            SEVERITY_ERROR,
            "No header row found in the first {0} rows. Expected columns: "
            "{1}.".format(_MAX_HEADER_SCAN_ROWS, wanted),
            sheet=sheet_name))
        return [], {}

    metadata = read_metadata(rows, header_index)
    mapping, complete = _column_index_map(
        rows[header_index], columns, sheet_name, issues)
    if not complete:
        return [], metadata

    key_attr = columns[0][0]
    objects = []
    for offset, row in enumerate(rows[header_index + 1:]):
        row_number = header_index + offset + 2      # 1-based, as Excel shows it
        if _row_is_blank(row):
            continue
        values = _read_row(row, mapping, columns, sheet_name, row_number, issues)
        if not values.get(key_attr):
            issues.append(Issue(
                SEVERITY_WARNING,
                "Row ignored: no {0}.".format(columns[0][1]),
                sheet=sheet_name, row=row_number))
            continue
        values["sheet"] = sheet_name
        values["source_row"] = row_number
        objects.append(factory(**values))
    return objects, metadata


def _match_sheets(grids):
    """``{canonical name: rows}`` -- sheet names folded, so spacing is forgiven."""
    folded = {}
    for name, rows in grids.items():
        folded.setdefault(fold(name), (name, rows))

    matched = {}
    present = []
    for canonical in models.ALL_SHEETS:
        names = ([fold(alias) for alias in models.INFO_SHEET_ALIASES]
                 if canonical == models.SHEET_INFO else [fold(canonical)])
        for name in names:
            found = folded.get(name)
            if found is not None:
                matched[canonical] = found[1]
                present.append(canonical)
                break
    return matched, tuple(present)


def _check_units(metadata, issues):
    """Millimetres, declared or assumed -- and never silently something else."""
    declared = None
    for key in ("units", "unit"):
        if key in metadata:
            declared = metadata[key]
            break
    if declared is None:
        issues.append(Issue(
            SEVERITY_WARNING,
            "No UNITS declared in the workbook. Reading every length as "
            "millimetres. Add an {0} sheet with a 'UNITS | mm' row to confirm "
            "it.".format(models.SHEET_INFO)))
        return "mm"
    if fold(declared) in ("mm", "millimetre", "millimetres", "millimeter",
                          "millimeters"):
        return "mm"
    issues.append(Issue(
        SEVERITY_ERROR,
        "Workbook declares UNITS as {0!r}. Only millimetres are supported — "
        "convert the schedule to mm and declare 'mm'.".format(declared)))
    return declared


def parse_grid(grids, path=None, mode=models.MODE_CREATE_ALL):
    """``(WorkbookData, issues)`` from raw sheet grids. No dependencies.

    Structural problems only -- a sheet that is absent, a column that is not
    there, a cell that is not a number. Whether the numbers make engineering
    sense is :mod:`validation`'s question, and keeping the two apart is what
    lets a workbook report "column missing" without also reporting four hundred
    consequences of it.

    Every sheet the parser knows is read whatever the mode, and *mode* decides
    only which absences are errors. Switching mode in the window then costs
    nothing: the workbook is read once, and a file with placement in it is ready
    to create structure the moment the user asks for that instead of rebar.
    """
    issues = []
    matched, present = _match_sheets(grids or {})
    needed = models.required_sheets(mode)

    for canonical in needed:
        if canonical not in matched:
            issues.append(Issue(
                SEVERITY_ERROR,
                "Required sheet {0!r} is missing from the workbook.".format(
                    canonical)))

    parsed = {}
    # The cover sheet wins where it says anything, because it is the sheet the
    # user maintains on purpose; a title block above a header is the fallback
    # for a schedule that arrives with one.
    metadata = read_key_value_sheet(matched.get(models.SHEET_INFO))
    for canonical, columns, factory in _SHEET_SPECS:
        rows = matched.get(canonical)
        if rows is None:
            parsed[canonical] = []
            continue
        objects, sheet_metadata = _parse_sheet(
            rows, columns, factory, canonical, issues)
        parsed[canonical] = objects
        for key, value in sheet_metadata.items():
            metadata.setdefault(key, value)

    for canonical in models.PLACEMENT_SHEETS:
        if canonical in matched and canonical not in needed:
            issues.append(Issue(
                SEVERITY_INFO,
                "Sheet {0!r} was read but is not used in this mode — the "
                "structure already exists, so it supplies its own "
                "geometry.".format(canonical)))

    units = _check_units(metadata, issues)

    data = models.WorkbookData(
        path=path,
        units=units,
        footing_types=parsed[models.SHEET_FOOTING_TYPES],
        column_types=parsed[models.SHEET_COLUMN_TYPES],
        footing_rebar=parsed[models.SHEET_FOOTING_REBAR],
        column_rebar=parsed[models.SHEET_COLUMN_REBAR],
        footing_placement=parsed[models.SHEET_FOOTING_PLACEMENT],
        column_placement=parsed[models.SHEET_COLUMN_PLACEMENT],
        metadata=metadata,
        sheets_present=present,
    )
    return data, issues


def load(path, mode=models.MODE_CREATE_ALL):
    """Read and parse in one step: ``(WorkbookData, issues)``.

    What the UI calls. Parsing is skipped when the file could not be opened at
    all, so the user gets the one message that matters instead of that message
    followed by four "sheet missing" ones.
    """
    grids, issues = read_grid(path)
    if models.has_errors(issues):
        return models.WorkbookData(path=path), issues
    data, parse_issues = parse_grid(grids, path=path, mode=mode)
    return data, issues + parse_issues
