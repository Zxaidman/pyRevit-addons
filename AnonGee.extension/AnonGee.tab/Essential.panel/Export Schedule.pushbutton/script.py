#! python3
# -*- coding: utf-8 -*-

import os
import clr
import System

clr.AddReference("RevitAPI")
clr.AddReference("RevitAPIUI")
from Autodesk.Revit.DB import (
    FilteredElementCollector, ViewSchedule, ViewScheduleExportOptions,
    ExportTextQualifier, ExportColumnHeaders, Transaction, SectionType
)

clr.AddReference("PresentationFramework")
clr.AddReference("PresentationCore")
clr.AddReference("WindowsBase")
clr.AddReference("System.Xaml")
clr.AddReference("System.Windows.Forms")

from System.Windows.Markup import XamlReader
from System.Windows.Media.Imaging import BitmapImage
from System.IO import FileStream, FileMode, FileAccess
from System.Windows.Controls import ListBoxItem, CheckBox, TextBlock
from System.Windows.Threading import Dispatcher, DispatcherPriority
from System import Action, Uri, UriKind
from System.Windows import Visibility, Thickness, MessageBox, MessageBoxButton, MessageBoxImage
import System.Windows.Forms as WinForms

doc       = __revit__.ActiveUIDocument.Document
LOCAL_DIR = os.path.dirname(__file__)


def _obj_arr(*items):
    """Build a .NET Object[] reliably. `System.Array[object]` (Python's
    builtin object) raises 'type expected' in this CPython3/pythonnet engine,
    so construct via CreateInstance instead."""
    arr = System.Array.CreateInstance(System.Object, len(items))
    for i, item in enumerate(items):
        arr[i] = item
    return arr

# ── stdlib with fallbacks (pyRevit CPython3 has limited stdlib path) ──────────
try:
    import traceback as _tb
    def _fmt_exc():
        return _tb.format_exc()
except ImportError:
    import sys
    def _fmt_exc():
        t, v, _ = sys.exc_info()
        return "{}: {}".format(t.__name__ if t else "Error", v)

try:
    import tempfile
    _TEMP_DIR = tempfile.gettempdir()
except ImportError:
    _TEMP_DIR = (os.environ.get("TEMP") or os.environ.get("TMP")
                 or os.path.expanduser("~"))

# ── AnonGee brand palette for Excel (hex, no alpha) — see Brand Guidelines §16.1
_BR_VIVID_RED   = "E02020"   # ColorVividRed   — column-header band
_BR_CHARCOAL    = "141414"   # ColorCharcoal   — title band
_BR_SILVER      = "C0C8D8"   # ColorSilverSteel — cell borders
_BR_WHITE       = "FFFFFF"   # ColorPureWhite  — header text, odd rows
_BR_OFFWHITE    = "F4F4F6"   # ColorOffWhite   — zebra (alternating) rows
_BR_CHARTEXT    = "141414"   # body data text (primary text)

# Primary brand face. Excel stores ONE font name per cell (no fallback chain
# like WPF) and substitution is host-controlled, so install JetBrains Mono on
# export hosts. Recommended manual fallback if unavailable: Cambria. (§16.1)
_BR_FONT        = "JetBrains Mono"  # Brand typeface — monospaced for tabular clarity
_BR_TITLE_SIZE  = 13         # H3-ish, bold
_BR_HEADER_SIZE = 10         # H4 / column headers, bold
_BR_BODY_SIZE   = 10         # body data

# Number formats — Indian digit grouping (lakh/crore). Excel repeats the
# leftmost 2-digit group, so "#,##,##0" yields 1,00,000 and 1,00,00,000.
_FMT_INT   = "#,##,##0"      # integers          -> 1,00,000
_FMT_FLOAT = "#,##,##0.000"  # floats, 3 decimals -> 1,00,000.000

_MIN_ROW_H = 20              # minimum Excel row height (ruler units)


# ── TSV parser (no csv module needed) ────────────────────────────────────────
def _parse_tsv_row(line):
    """Parse one tab-delimited line, honouring double-quoted fields."""
    fields, buf, in_q = [], [], False
    i = 0
    while i < len(line):
        ch = line[i]
        if ch == '"':
            if in_q and i + 1 < len(line) and line[i + 1] == '"':
                buf.append('"'); i += 2; continue
            in_q = not in_q
        elif ch == '\t' and not in_q:
            fields.append(''.join(buf)); buf = []
        else:
            buf.append(ch)
        i += 1
    fields.append(''.join(buf))
    return fields


def _read_tsv(path):
    for enc in ('utf-8-sig', 'cp1252', 'latin-1'):
        try:
            rows = []
            with open(path, 'r', encoding=enc) as fh:
                for line in fh:
                    rows.append(_parse_tsv_row(line.rstrip('\r\n')))
            return rows
        except (UnicodeDecodeError, Exception):
            continue
    return []


# ── Unit stripper (no re module needed) ──────────────────────────────────────
def _norm_unit(u):
    """Normalize a unit token: lowercase, map superscripts (² ³) and the
    'Â' UTF-8/latin-1 mojibake to plain digits, then drop all spaces."""
    u = u.strip().lower()
    u = (u.replace(u"\xb2", "2").replace(u"\xb3", "3")   # ² ³
           .replace(u"\xe2", "").replace(u"\xc2", ""))   # â Â artifacts
    return u.replace(" ", "")


_KNOWN_UNITS = set(_norm_unit(u) for u in [
    # length
    "mm", "cm", "m", "km",
    # area
    "mm2", "cm2", "m2", "km2", "sq m", "sqm", u"mm²", u"cm²", u"m²",
    # volume
    "mm3", "cm3", "m3", "cu m", "cum", u"mm³", u"cm³", u"m³",
    # mass
    "g", "kg", "t", "ton", "tonne",
    # mass per length / area
    "kg/m", "kg/m2", "kg/m3", "kg/cm2", u"kg/m²", u"kg/m³",
    # force / pressure
    "n", "kn", "kn/m", "kn/m2", "n/mm2", u"n/mm²", "pa", "kpa", "mpa",
    # misc
    "l", "ml", "deg", u"°", "%",
])


def _strip_unit(value):
    """Convert 'number + known unit' (or plain numeric text) to int/float.
    Handles 'cm³', 'kg/m', superscripts and 'Â' artifacts. Leaves anything
    that is not a recognised unit pattern untouched (codes, ranges, text)."""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s:
        return value

    # Extract leading numeric prefix: optional sign, digits, one dot,
    # thousands commas.
    i = 0
    sign = ""
    if s[0] in "+-":
        sign = s[0]
        i = 1
    num = []
    seen_dot = False
    while i < len(s):
        ch = s[i]
        if ch.isdigit():
            num.append(ch)
        elif ch == "." and not seen_dot:
            num.append(ch)
            seen_dot = True
        elif ch == ",":
            pass  # thousands separator
        else:
            break
        i += 1

    if not num:
        return value  # no leading number → text, leave alone

    rest = s[i:]
    if rest and _norm_unit(rest) not in _KNOWN_UNITS:
        return value  # trailing text isn't a unit (e.g. "12B", "10-20")

    try:
        f = float(sign + "".join(num))
    except ValueError:
        return value
    return int(f) if f == int(f) else f


# ──────────────────────────────────────────────────────────────────────────────
# Dialog
# ──────────────────────────────────────────────────────────────────────────────
class ExportScheduleDialog(object):

    def __init__(self, schedules):
        self._schedules = schedules
        xaml_path = os.path.join(LOCAL_DIR, "ui.xaml")
        stream = FileStream(xaml_path, FileMode.Open, FileAccess.Read)
        try:
            self.window = XamlReader.Load(stream)
        finally:
            stream.Close()
        self._set_window_icon()
        self._bind()
        self._populate()

    def _set_window_icon(self):
        icon_path = os.path.join(LOCAL_DIR, "icon.png")
        if os.path.exists(icon_path):
            self.window.Icon = BitmapImage(Uri(icon_path, UriKind.Absolute))

    def _bind(self):
        w = self.window
        self._live_count         = w.FindName("LiveCount")
        self._search_box         = w.FindName("SearchBox")
        self._search_placeholder = w.FindName("SearchPlaceholder")
        self._schedule_list      = w.FindName("ScheduleList")
        self._chk_strip          = w.FindName("ChkStripUnits")
        self._chk_hidden         = w.FindName("ChkIncludeHidden")
        self._chk_zebra          = w.FindName("ChkZebraShading")
        self._chk_pdf            = w.FindName("ChkSavePdf")
        self._btn_all            = w.FindName("BtnAll")
        self._btn_none           = w.FindName("BtnNone")
        self._btn_export         = w.FindName("BtnExport")
        self._btn_close          = w.FindName("BtnClose")
        self._badge_info         = w.FindName("BadgeInfo")
        self._badge_success      = w.FindName("BadgeSuccess")
        self._badge_error        = w.FindName("BadgeError")
        self._info_text          = w.FindName("InfoText")
        self._success_text       = w.FindName("SuccessText")
        self._error_text         = w.FindName("ErrorText")

        self._search_box.TextChanged += self._on_search_changed
        self._search_box.GotFocus    += self._on_search_focus
        self._search_box.LostFocus   += self._on_search_blur
        self._btn_all.Click          += self._on_all
        self._btn_none.Click         += self._on_none
        self._btn_export.Click       += self._on_export
        self._btn_close.Click        += lambda s, e: self.window.Close()

    def _populate(self):
        n = len(self._schedules)
        self._live_count.Text = "{} schedule{}".format(n, "s" if n != 1 else "")
        for sched in self._schedules:
            tb = TextBlock()
            tb.Text = sched.Name
            chk = CheckBox()
            chk.Content   = tb
            chk.IsChecked = True
            chk.Margin    = Thickness(2)
            item = ListBoxItem()
            item.Content = chk
            item.Tag     = sched
            self._schedule_list.Items.Add(item)
        self._show_info("{} schedule{} loaded — check and export".format(
            n, "s" if n != 1 else ""))

    # ── Search ──
    def _on_search_focus(self, sender, args):
        self._search_placeholder.Visibility = Visibility.Collapsed

    def _on_search_blur(self, sender, args):
        if not self._search_box.Text:
            self._search_placeholder.Visibility = Visibility.Visible

    def _on_search_changed(self, sender, args):
        q = (self._search_box.Text or "").lower()
        for item in self._schedule_list.Items:
            item.Visibility = (
                Visibility.Visible if q in item.Content.Content.Text.lower()
                else Visibility.Collapsed)

    # ── All / None ──
    def _on_all(self, sender, args):
        for item in self._schedule_list.Items:
            item.Content.IsChecked = True

    def _on_none(self, sender, args):
        for item in self._schedule_list.Items:
            item.Content.IsChecked = False

    # ── Badges ──
    def _show_info(self, msg):
        self._info_text.Text           = msg
        self._badge_info.Visibility    = Visibility.Visible
        self._badge_success.Visibility = Visibility.Collapsed
        self._badge_error.Visibility   = Visibility.Collapsed

    def _show_success(self, msg):
        self._success_text.Text        = msg
        self._badge_info.Visibility    = Visibility.Collapsed
        self._badge_success.Visibility = Visibility.Visible
        self._badge_error.Visibility   = Visibility.Collapsed

    def _show_error(self, msg):
        self._error_text.Text          = msg
        self._badge_info.Visibility    = Visibility.Collapsed
        self._badge_success.Visibility = Visibility.Collapsed
        self._badge_error.Visibility   = Visibility.Visible

    def _flush_ui(self):
        Dispatcher.CurrentDispatcher.Invoke(
            Action(lambda: None), DispatcherPriority.Background)

    # ── Export button ──
    def _on_export(self, sender, args):
        try:
            self._do_export()
        except Exception:
            tb = _fmt_exc().strip().splitlines()
            self._show_error(tb[-1] if tb else "Unknown error")
            self._btn_export.IsEnabled = True

    def _do_export(self):
        selected = [item.Tag for item in self._schedule_list.Items
                    if item.Content.IsChecked]
        if not selected:
            self._show_error("Select at least one schedule.")
            return

        dlg            = WinForms.SaveFileDialog()
        dlg.Title      = "Export Schedules — Select Save Location"
        dlg.Filter     = "Excel Workbook (*.xlsx)|*.xlsx"
        dlg.DefaultExt = "xlsx"
        dlg.FileName   = "{}_Schedules.xlsx".format(doc.Title)
        if dlg.ShowDialog() != WinForms.DialogResult.OK:
            return

        save_path      = dlg.FileName
        strip_units    = bool(self._chk_strip.IsChecked)
        include_hidden = bool(self._chk_hidden.IsChecked)
        zebra_shading  = bool(self._chk_zebra.IsChecked)
        save_pdf       = bool(self._chk_pdf.IsChecked)

        self._show_info("Exporting {} schedule(s)...".format(len(selected)))
        self._btn_export.IsEnabled = False
        self._flush_ui()

        ok, message = export_schedules_to_excel(
            selected, save_path, strip_units, include_hidden, zebra_shading, save_pdf)

        if ok:
            self._show_success(message)
            self._flush_ui()
            try:
                os.startfile(save_path)
            except Exception:
                pass
            # If a PDF was requested but skipped, keep the window open so the
            # reason stays visible; otherwise auto-close on clean success.
            if not (save_pdf and "PDF skipped" in message):
                self.window.Close()
            else:
                self._btn_export.IsEnabled = True
        else:
            lines = [l.strip() for l in message.strip().splitlines() if l.strip()]
            self._show_error(lines[-1] if lines else message)
            self._btn_export.IsEnabled = True

    def show(self):
        self.window.ShowDialog()


# ──────────────────────────────────────────────────────────────────────────────
# Excel export — openpyxl only, no COM, no Excel installation required
# ──────────────────────────────────────────────────────────────────────────────
def export_schedules_to_excel(schedules, output_path,
                               strip_units=True, include_hidden=False,
                               shade_rows=True, save_pdf=False):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError as e:
        return False, "openpyxl not available: {}".format(e)

    try:
        wb = Workbook()
        wb.remove(wb.active)   # drop the default blank sheet

        # ── Brand styles ──────────────────────────────────────────────────
        thin = Side(style="thin", color=_BR_SILVER)        # Silver Steel grid
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title band — Charcoal Black, white bold, centred both ways
        title_fill  = PatternFill(fill_type="solid", fgColor=_BR_CHARCOAL)
        title_font  = Font(name=_BR_FONT, size=_BR_TITLE_SIZE, bold=True, color=_BR_WHITE)
        title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column headers — Vivid Red, white bold, centred both ways
        hdr_fill  = PatternFill(fill_type="solid", fgColor=_BR_VIVID_RED)
        hdr_font  = Font(name=_BR_FONT, size=_BR_HEADER_SIZE, bold=True, color=_BR_WHITE)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Body — Charcoal text, vertical centre; horizontal set per cell type
        body_font   = Font(name=_BR_FONT, size=_BR_BODY_SIZE, color=_BR_CHARTEXT)
        left_align  = Alignment(horizontal="left",  vertical="center", wrap_text=False)
        right_align = Alignment(horizontal="right", vertical="center", wrap_text=False)

        zebra_fill = PatternFill(fill_type="solid", fgColor=_BR_OFFWHITE)
        white_fill = PatternFill(fill_type="solid", fgColor=_BR_WHITE)

        for i, sched in enumerate(schedules):
            # ── Export Revit schedule to temp TSV ─────────────────────────
            fn        = "ag_s{}.txt".format(i)
            temp_path = os.path.join(_TEMP_DIR, fn)

            opt                = ViewScheduleExportOptions()
            opt.FieldDelimiter = "\t"
            opt.TextQualifier  = ExportTextQualifier.DoubleQuote
            opt.ColumnHeaders  = ExportColumnHeaders.OneRow

            hidden_fields = []
            if include_hidden:
                defn = sched.Definition
                t_un = Transaction(doc, "Temp Unhide Fields")
                t_un.Start()
                for j in range(defn.GetFieldCount()):
                    fld = defn.GetField(j)
                    if fld.IsHidden:
                        hidden_fields.append(j)
                        fld.IsHidden = False
                t_un.Commit()

            sched.Export(_TEMP_DIR, fn, opt)

            if include_hidden and hidden_fields:
                t_re = Transaction(doc, "Restore Hidden Fields")
                t_re.Start()
                for j in hidden_fields:
                    sched.Definition.GetField(j).IsHidden = True
                t_re.Commit()

            # ── Read TSV (inline parser — no stdlib csv needed) ───────────
            rows = _read_tsv(temp_path)
            try:
                os.remove(temp_path)
            except Exception:
                pass
            if not rows:
                continue

            row_count = len(rows)
            col_count = max(len(r) for r in rows)

            # ── Create worksheet ──────────────────────────────────────────
            sn = sched.Name
            for ch in [':', '\\', '/', '?', '*', '[', ']']:
                sn = sn.replace(ch, '_')
            ws = wb.create_sheet(title=sn[:31])

            # ── Write all cell values ─────────────────────────────────────
            for ri, row_data in enumerate(rows, 1):
                for ci in range(1, col_count + 1):
                    ws.cell(row=ri, column=ci,
                            value=row_data[ci - 1] if ci - 1 < len(row_data) else "")

            # ── Detect header depth ───────────────────────────────────────
            # Pattern A: R1=title, R2=blank, R3=col headers  → h_depth=3
            # Pattern B: R1=title, R2=col headers            → h_depth=2
            h_depth = 2
            if row_count >= 3:
                r2_blank    = all(not str(v).strip() for v in rows[1])
                r3_has_data = any(str(v).strip() for v in rows[2])
                if r2_blank and r3_has_data:
                    h_depth = 3

            # ── Strip units FIRST so numeric cells become real numbers and
            #    can be detected for right-alignment below ─────────────────
            if strip_units:
                for r in range(h_depth + 1, row_count + 1):
                    for c in range(1, col_count + 1):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell.value, str):
                            cell.value = _strip_unit(cell.value)

            # ── Title band (rows above the column-header row): Charcoal ───
            for r in range(1, h_depth):
                for c in range(1, col_count + 1):
                    cell           = ws.cell(row=r, column=c)
                    cell.fill      = title_fill
                    cell.font      = title_font
                    cell.alignment = title_align
                    cell.border    = bdr
            # Merge the title (row 1) across all columns
            if col_count > 1:
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1,   end_column=col_count)

            # ── Column header row (row h_depth): Vivid Red ────────────────
            for c in range(1, col_count + 1):
                cell           = ws.cell(row=h_depth, column=c)
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = hdr_align
                cell.border    = bdr

            # ── Body rows: zebra fill, Charcoal text, vertical centre.
            #    Numbers → right-aligned + Indian grouping; text → left.
            #    Integers show as 0, floats as 0.000 (e.g. 1,00,000.000).
            for r in range(h_depth + 1, row_count + 1):
                is_even = (r - h_depth) % 2 == 0
                for c in range(1, col_count + 1):
                    cell           = ws.cell(row=r, column=c)
                    cell.font      = body_font
                    cell.border    = bdr
                    cell.fill      = (zebra_fill if shade_rows and is_even
                                      else white_fill)
                    val = cell.value
                    if isinstance(val, bool):
                        cell.alignment = left_align
                    elif isinstance(val, int):
                        cell.number_format = _FMT_INT      # 0  -> 1,00,000
                        cell.alignment     = right_align
                    elif isinstance(val, float):
                        cell.number_format = _FMT_FLOAT    # 0.000 -> 1,00,000.000
                        cell.alignment     = right_align
                    else:
                        cell.alignment = left_align

            # ── Row heights — minimum 20 for every row ────────────────────
            ws.row_dimensions[1].height = max(24, _MIN_ROW_H)   # title
            if h_depth == 3:
                ws.row_dimensions[2].height = _MIN_ROW_H        # separator
            ws.row_dimensions[h_depth].height = max(30, _MIN_ROW_H)  # headers (wrap)
            for r in range(h_depth + 1, row_count + 1):
                ws.row_dimensions[r].height = _MIN_ROW_H        # body

            # ── Column width — TIGHT estimate, no extra padding ───────────
            # Fallback only: if Excel is available, _excel_polish() runs a
            # true Columns.AutoFit afterward and overwrites these. Excludes
            # the merged title row (row 1) so a long schedule name can't
            # inflate column A.
            for c in range(1, col_count + 1):
                ltr     = get_column_letter(c)
                longest = 0
                for r in range(2, min(row_count + 1, 200)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        longest = max(longest, len(str(v)))
                # Body font is 9 (< Excel's default 11), so chars need
                # slightly less than 1 width-unit each. No extra padding.
                width = longest * 0.92 + 0.7
                ws.column_dimensions[ltr].width = max(5, min(width, 45))

            # ── Page setup: landscape, fit all columns on one page wide ───
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth  = 1      # all columns on one page wide
            ws.page_setup.fitToHeight = 0      # as many pages tall as needed
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

            # ── Freeze panes below header row ─────────────────────────────
            ws.freeze_panes = ws.cell(row=h_depth + 1, column=1)

        wb.save(output_path)

        msg = "Exported {} schedule(s) to {}".format(
            len(schedules), os.path.basename(output_path))

        # Excel polish pass: true column AutoFit + page setup, and PDF if
        # requested. Skipped gracefully if Excel is not installed (the xlsx
        # already has tight estimated widths + landscape/fit page setup).
        polished, pdf_ok, info = _excel_polish(output_path, save_pdf)
        if save_pdf and not pdf_ok:
            msg += " — PDF skipped: {}".format(info)

        return True, msg

    except Exception:
        return False, _fmt_exc()


def _excel_polish(xlsx_path, want_pdf):
    """Open the saved xlsx in Excel via COM to apply TRUE column AutoFit and
    page setup (landscape, fit to one page wide), then optionally export PDF.

    Returns (polished, pdf_ok, info). Degrades gracefully without Excel — the
    openpyxl file already carries tight widths and landscape/fit page setup.
    """
    try:
        from System import Type, Activator, Reflection
    except Exception as e:
        return False, False, "interop unavailable ({})".format(e)

    fg = Reflection.BindingFlags.GetProperty
    fs = Reflection.BindingFlags.SetProperty
    fi = Reflection.BindingFlags.InvokeMethod

    xt = Type.GetTypeFromProgID("Excel.Application")
    if xt is None:
        return False, False, "Microsoft Excel not installed"

    xa = None
    try:
        xa = Activator.CreateInstance(xt)
        xt.InvokeMember("Visible",       fs, None, xa, _obj_arr(False))
        xt.InvokeMember("DisplayAlerts", fs, None, xa, _obj_arr(False))

        wbs = xt.InvokeMember("Workbooks", fg, None, xa, None)
        wb  = wbs.GetType().InvokeMember("Open", fi, None, wbs, _obj_arr(xlsx_path))
        wbt = wb.GetType()

        sheets = wbt.InvokeMember("Worksheets", fg, None, wb, None)
        st     = sheets.GetType()
        count  = st.InvokeMember("Count", fg, None, sheets, None)

        for idx in range(1, count + 1):
            ws  = st.InvokeMember("Item", fg, None, sheets, _obj_arr(idx))
            wst = ws.GetType()

            # True column AutoFit — uses the real cell font, matching what you
            # get by double-clicking the column border in Excel. Row heights
            # are left as set (min 20) — we deliberately do NOT AutoFit rows,
            # which would shrink them below the 20-unit minimum.
            cols = wst.InvokeMember("Columns", fg, None, ws, None)
            cols.GetType().InvokeMember("AutoFit", fi, None, cols, None)

            # Page setup
            ps  = wst.InvokeMember("PageSetup", fg, None, ws, None)
            pst = ps.GetType()
            pst.InvokeMember("Orientation",    fs, None, ps, _obj_arr(2))      # xlLandscape
            pst.InvokeMember("Zoom",           fs, None, ps, _obj_arr(False))
            pst.InvokeMember("FitToPagesWide", fs, None, ps, _obj_arr(1))
            pst.InvokeMember("FitToPagesTall", fs, None, ps, _obj_arr(False))

        wbt.InvokeMember("Save", fi, None, wb, None)   # persist true widths

        pdf_ok = False
        if want_pdf:
            pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
            wbt.InvokeMember("ExportAsFixedFormat", fi, None, wb,
                             _obj_arr(0, pdf_path))     # 0 = xlTypePDF
            pdf_ok = True

        wbt.InvokeMember("Close", fi, None, wb, _obj_arr(False))
        return True, pdf_ok, "ok"

    except Exception:
        tb = _fmt_exc().strip().splitlines()
        return False, False, (tb[-1] if tb else "unknown error")
    finally:
        if xa is not None:
            try:
                xt.InvokeMember("Quit", fi, None, xa, None)
                System.Runtime.InteropServices.Marshal.ReleaseComObject(xa)
            except Exception:
                pass


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────
def run():
    all_schedules = list(FilteredElementCollector(doc).OfClass(ViewSchedule).ToElements())

    exportable = []
    for s in all_schedules:
        if s.IsTemplate:
            continue
        try:
            if s.GetTableData().GetSectionData(SectionType.Body).NumberOfRows > 3:
                exportable.append(s)
        except Exception:
            continue

    exportable.sort(key=lambda s: s.Name)

    if not exportable:
        MessageBox.Show(
            "No exportable schedules found in this document.\n\n"
            "Schedules must have more than 3 data rows and must not be templates.",
            "Export Schedule",
            MessageBoxButton.OK,
            MessageBoxImage.Warning)
        return

    ExportScheduleDialog(exportable).show()


if __name__ == "__main__":
    # Module-level guard: any unhandled exception shows a dialog instead of
    # crashing the CPython3 engine (which would require a Revit restart).
    try:
        run()
    except Exception:
        MessageBox.Show(
            _fmt_exc()[-3000:],
            "Export Schedule — Unhandled Error",
            MessageBoxButton.OK,
            MessageBoxImage.Error)
