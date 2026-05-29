# -*- coding: utf-8 -*-
"""
output/excel_writer.py
Builds the BBS Excel workbook using openpyxl.
Includes dynamic Excel formulas for totals and Auto-Fit columns.
"""

import os
from collections import defaultdict

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

C_HEADER_BG   = "2C3E50"
C_HEADER_FG   = "FFFFFF"
C_FLOOR_BG    = "2980B9"
C_FLOOR_FG    = "FFFFFF"
C_MEMBER_BG   = "EBF5FB"
C_MEMBER_FG   = "1A5276"
C_COL_HEAD_BG = "3498DB"
C_COL_HEAD_FG = "FFFFFF"
C_SUBTOTAL_BG = "D5D8DC"
C_GRAND_BG    = "BDC3C7"
C_ROW_ALT     = "F5F5F5"
C_ROW_PLAIN   = "FFFFFF"
C_NEW         = "D5F5E3"
C_CHANGED     = "FDEBD0"
C_DELETED     = "FADBD8"
C_CALC_BG     = "F0F3F4"

_thin  = Side(style="thin",   color="BDC3C7")
_thick = Side(style="medium", color="7F8C8D")
BORDER_THIN  = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
BORDER_THICK = Border(left=_thick, right=_thick, top=_thick, bottom=_thick)

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, size=10, color="000000", italic=False): return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_row(ws, row, values, bold=False, bg=None, fg="000000", aligns=None, number_formats=None, border=BORDER_THIN):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _font(bold=bold, color=fg)
        if bg: cell.fill = _fill(bg)
        cell.border = border
        if aligns and col_idx - 1 < len(aligns): cell.alignment = _align(aligns[col_idx - 1])
        else: cell.alignment = _align("center" if col_idx > 4 else "left")
        if number_formats and col_idx - 1 < len(number_formats):
            fmt = number_formats[col_idx - 1]
            if fmt: cell.number_format = fmt

# Column definition: (header, width, alignment, number_format)
# Column positions (1-based, used by _build_bbs_row for dynamic Excel formulas):
#   1:#  2:BarMark  3:BarType  4:Detail  5:Dia  6:Shape
#   7:A  8:B  9:C  10:D  11:E  12:No.Bars  13:Cut.Len  14:Tot.Len  [15:UnitWt]  last:TotWt
BBS_COLUMNS = [
    ("#",               5,  "center", None),       # 1
    ("Bar Mark",        10, "center", None),        # 2
    ("Bar Type",        14, "left",   None),        # 3
    ("Detail",          18, "left",   None),        # 4
    ("Dia (mm)",         9, "center", None),        # 5
    ("Shape Code",       9, "center", None),        # 6
    ("A (mm)",           9, "center", "#,##0"),     # 7
    ("B (mm)",           9, "center", "#,##0"),     # 8
    ("C (mm)",           9, "center", "#,##0"),     # 9
    ("D (mm)",           9, "center", "#,##0"),     # 10
    ("E (mm)",           9, "center", "#,##0"),     # 11  ← NEW
    ("No. of Bars",      9, "center", "#,##0"),     # 12
    ("Cut. Length (mm)", 14,"center", "#,##0"),     # 13
    ("Total Length (m)", 14,"center", "#,##0.00"),  # 14
    ("Unit Wt (kg/m)",   12,"center", "#,##0.000"), # 15  (toggled)
    ("Total Wt (kg)",    12,"center", "#,##0.000"), # 16 or 15
]

BBS_COLUMNS_NO_UW = [c for c in BBS_COLUMNS if c[0] != "Unit Wt (kg/m)"]

def _auto_fit_columns(ws, min_width=8, max_width=60):
    """
    Adjusts column widths based on cell content.
    Skips MergedCell objects (they have no column_letter attribute).
    Applies a sensible min/max clamp.
    """
    from openpyxl.cell.cell import MergedCell
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # MergedCell instances have no column_letter — skip them
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None:
                continue
            col = cell.column_letter
            try:
                cell_len = len(str(cell.value))
            except Exception:
                cell_len = 8
            col_widths[col] = max(col_widths.get(col, min_width), cell_len + 2)

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = min(max(width, min_width), max_width)

def _set_row_heights(ws, default_height=20):
    """Sets all populated row heights to default_height (points)."""
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = default_height


def write_bbs_workbook(records, output_path, project_info, show_unit_weight=True, include_calc_sheet=True, include_summary_sheet=True, revision_changes=None, standard_module=None, progress_callback=None):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    cols = BBS_COLUMNS if show_unit_weight else BBS_COLUMNS_NO_UW

    if progress_callback: progress_callback(0, 3, "Writing BBS sheet…")
    ws_bbs = wb.create_sheet("BBS")
    _write_bbs_sheet(ws_bbs, records, project_info, cols, revision_changes, show_unit_weight)
    _auto_fit_columns(ws_bbs)
    _set_row_heights(ws_bbs, 20)

    if include_calc_sheet:
        if progress_callback: progress_callback(1, 3, "Writing Calculation sheet…")
        ws_calc = wb.create_sheet("Calculation")
        _write_calc_sheet(ws_calc, records, project_info, standard_module)
        _auto_fit_columns(ws_calc)
        _set_row_heights(ws_calc, 20)

    if include_summary_sheet:
        if progress_callback: progress_callback(2, 3, "Writing Summary sheet…")
        ws_sum = wb.create_sheet("Summary")
        _write_summary_sheet(ws_sum, records, project_info)
        _auto_fit_columns(ws_sum)
        _set_row_heights(ws_sum, 20)

    wb.save(output_path)
    if progress_callback: progress_callback(3, 3, "Saved ✓")
    return output_path

def _write_bbs_sheet(ws, records, project_info, cols, revision_changes, show_unit_weight):
    n_cols = len(cols)
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=f"BAR BENDING SCHEDULE — {project_info.get('project_name','')}")
    c.font = _font(bold=True, size=13, color=C_HEADER_FG)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _align("center")

    row = 2
    meta = [
        f"Drawing No: {project_info.get('drawing_no','')}",
        f"Prepared By: {project_info.get('prepared_by','')}",
        f"Date: {project_info.get('date','')}",
        f"Standard: {project_info.get('standard','')}",
        f"Rev: {project_info.get('revision','A')}",
    ]
    per_block = max(1, n_cols // len(meta))
    for i, txt in enumerate(meta):
        c1 = i * per_block + 1
        c2 = min(c1 + per_block - 1, n_cols)
        if c1 > n_cols: break
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=txt)
        cell.font = _font(size=9, color="555555")
        cell.alignment = _align("left")

    row = 4
    headers  = [c[0] for c in cols]
    aligns_h = [c[2] for c in cols]
    _set_row(ws, row, headers, bold=True, bg=C_COL_HEAD_BG, fg=C_COL_HEAD_FG, aligns=aligns_h, border=BORDER_THICK)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = ws.cell(row=5, column=1)

    grouped = defaultdict(lambda: defaultdict(list))
    for r in records: grouped[r.floor_level][r.member_name].append(r)

    data_row = 5
    serial = 0
    len_col_let = get_column_letter(n_cols - (2 if show_unit_weight else 1))
    wt_col_let  = get_column_letter(n_cols)

    for floor_level in sorted(grouped.keys()):
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=n_cols)
        c = ws.cell(row=data_row, column=1, value=f"  LEVEL: {floor_level.upper()}")
        c.font, c.fill, c.alignment = _font(bold=True, size=11, color=C_FLOOR_FG), _fill(C_FLOOR_BG), _align("left")
        ws.row_dimensions[data_row].height = 18
        data_row += 1
        floor_start = data_row

        for member_name in sorted(grouped[floor_level].keys()):
            member_records = grouped[floor_level][member_name]
            ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=n_cols)
            mtype = member_records[0].member_type if member_records else ""
            c = ws.cell(row=data_row, column=1, value=f"    {mtype}: {member_name}")
            c.font, c.fill, c.alignment = _font(bold=True, size=10, color=C_MEMBER_FG), _fill(C_MEMBER_BG), _align("left")
            data_row += 1

            member_start = data_row
            for rec_idx, r in enumerate(member_records):
                serial += 1
                bg = C_ROW_ALT if rec_idx % 2 == 1 else C_ROW_PLAIN
                rev_key = f"{r.bar_mark}|{r.member_name}|{r.floor_level}"
                if revision_changes:
                    state = revision_changes.get(rev_key, "")
                    if state == "new": bg = C_NEW
                    elif state == "changed": bg = C_CHANGED
                    elif state == "deleted": bg = C_DELETED

                row_vals = _build_bbs_row(serial, r, show_unit_weight, data_row, cols)
                num_fmts = [c[3] for c in cols]
                _set_row(ws, data_row, row_vals, bg=bg, number_formats=num_fmts)
                data_row += 1

            member_end = data_row - 1
            sub_vals = [""] * n_cols
            sub_vals[0] = f"Subtotal — {member_name}"
            sub_vals[n_cols - (3 if show_unit_weight else 2)] = f"=SUM({len_col_let}{member_start}:{len_col_let}{member_end})"
            sub_vals[n_cols - 1] = f"=SUM({wt_col_let}{member_start}:{wt_col_let}{member_end})"
            _set_row(ws, data_row, sub_vals, bold=True, bg=C_SUBTOTAL_BG, number_formats=[None] * (n_cols - 2) + ["#,##0.00", "#,##0.000"])
            data_row += 1

        floor_end = data_row - 1
        gt_vals = [""] * n_cols
        gt_vals[0] = f"Floor Total — {floor_level}"
        gt_vals[n_cols - (3 if show_unit_weight else 2)] = f"=SUM({len_col_let}{floor_start}:{len_col_let}{floor_end})/2"
        gt_vals[n_cols - 1] = f"=SUM({wt_col_let}{floor_start}:{wt_col_let}{floor_end})/2"
        _set_row(ws, data_row, gt_vals, bold=True, bg=C_GRAND_BG, number_formats=[None] * (n_cols - 2) + ["#,##0.00", "#,##0.000"])
        data_row += 2

    data_row += 1
    ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=n_cols)
    c = ws.cell(row=data_row, column=1, value="GRAND TOTAL BY DIAMETER (STATIC SUMMARY)")
    c.font, c.fill, c.alignment = _font(bold=True, size=11, color=C_HEADER_FG), _fill(C_HEADER_BG), _align("center")
    data_row += 1

    dia_totals = defaultdict(float)
    dia_lengths = defaultdict(float)
    for r in records:
        dia_totals[r.diameter_mm]  += r.total_weight_kg
        dia_lengths[r.diameter_mm] += r.total_length_mm / 1000.0

    _set_row(ws, data_row, ["Diameter", "Total Length (m)", "Total Weight (kg)"] + [""] * (n_cols - 3), bold=True, bg=C_COL_HEAD_BG, fg=C_COL_HEAD_FG)
    data_row += 1
    for dia in sorted(dia_totals.keys()):
        _set_row(ws, data_row, [f"T{dia}", dia_lengths[dia], dia_totals[dia]] + [""] * (n_cols - 3), number_formats=[None, "#,##0.00", "#,##0.000"] + [None] * (n_cols - 3))
        data_row += 1

    overall_wt  = sum(r.total_weight_kg for r in records)
    overall_len = sum(r.total_length_mm / 1000.0 for r in records)
    _set_row(ws, data_row, ["TOTAL", overall_len, overall_wt] + [""] * (n_cols - 3), bold=True, bg=C_GRAND_BG, number_formats=[None, "#,##0.00", "#,##0.000"] + [None] * (n_cols - 3))

def _build_bbs_row(serial, r, show_unit_weight, row_idx, cols):
    """
    Build one BBS data row with dynamic Excel formula column references.

    Column layout (1-based Excel columns, matches BBS_COLUMNS above):
      1:#  2:BarMark  3:BarType  4:Detail  5:Dia  6:Shape
      7:A  8:B  9:C  10:D  11:E  12:No.Bars  13:Cut.Len  14:Tot.Len
      [15:UnitWt]  16-or-15:TotWt

    Formula columns use get_column_letter() so they are always correct
    regardless of column insertion/removal.
    """
    # Fixed column positions — match BBS_COLUMNS definition exactly
    COL_NO  = 12   # "No. of Bars"
    COL_CUT = 13   # "Cut. Length (mm)"
    COL_LEN = 14   # "Total Length (m)"
    COL_UW  = 15   # "Unit Wt (kg/m)"  — only when show_unit_weight=True
    COL_WT_UW  = 16  # "Total Wt (kg)" with unit weight col present
    COL_WT_NUW = 15  # "Total Wt (kg)" without unit weight col

    no_col  = get_column_letter(COL_NO)
    cut_col = get_column_letter(COL_CUT)
    len_col = get_column_letter(COL_LEN)

    # Total Length formula: No.Bars * Cut.Length / 1000  (mm → m)
    total_len_formula = f"={no_col}{row_idx}*{cut_col}{row_idx}/1000"

    # Total Weight formula: Tot.Length(m) * unit_weight
    if show_unit_weight:
        uw_col           = get_column_letter(COL_UW)
        total_wt_formula = f"={len_col}{row_idx}*{uw_col}{row_idx}"
    else:
        # Unit weight embedded as a literal — no separate column
        uw               = r.unit_weight_kg_per_m
        total_wt_formula = f"={len_col}{row_idx}*{uw}"

    row = [
        serial,                          # 1  #
        r.bar_mark,                      # 2  Bar Mark
        r.bar_type_name,                 # 3  Bar Type
        r.bar_detail,                    # 4  Detail
        r.diameter_mm,                   # 5  Dia (mm)
        r.shape_code,                    # 6  Shape Code
        r.dimensions.get("A", ""),       # 7  A (mm)
        r.dimensions.get("B", ""),       # 8  B (mm)
        r.dimensions.get("C", ""),       # 9  C (mm)
        r.dimensions.get("D", ""),       # 10 D (mm)
        r.dimensions.get("E", ""),       # 11 E (mm)  ← NEW
        r.no_of_bars,                    # 12 No. of Bars
        r.cutting_length_mm,             # 13 Cut. Length (mm)
        total_len_formula,               # 14 Total Length (m)
    ]
    if show_unit_weight:
        row.append(r.unit_weight_kg_per_m)   # 15 Unit Wt (kg/m)
    row.append(total_wt_formula)             # 16/15 Total Wt (kg)
    return row

def _write_calc_sheet(ws, records, project_info, standard_module):
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"CUTTING LENGTH CALCULATION SHEET — {project_info.get('project_name','')}")
    c.font, c.fill, c.alignment = _font(bold=True, size=12, color=C_HEADER_FG), _fill(C_HEADER_BG), _align("center")
    ws.row_dimensions[1].height = 20
    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1, value=f"Standard: {project_info.get('standard','')}   |   Date: {project_info.get('date','')}   |   Prepared By: {project_info.get('prepared_by','')}")
    ws.row_dimensions[2].height = 14

    headers = ["Bar Mark", "Member", "Dia (mm)", "Shape Code", "Formula / Calculation", "Bend Deduction (mm)"]
    _set_row(ws, 3, headers, bold=True, bg=C_COL_HEAD_BG, fg=C_COL_HEAD_FG, border=BORDER_THICK)
    ws.row_dimensions[3].height = 16
    ws.freeze_panes = ws.cell(row=4, column=1)

    sorted_records = sorted(records, key=lambda r: (r.floor_level, r.member_name, r.bar_mark))
    current_floor = None

    calc_row = 4
    for idx, r in enumerate(sorted_records):
        if r.floor_level != current_floor:
            current_floor = r.floor_level
            ws.merge_cells(start_row=calc_row, start_column=1, end_row=calc_row, end_column=6)
            c = ws.cell(row=calc_row, column=1, value=f"  LEVEL: {current_floor.upper()}")
            c.font, c.fill, c.alignment = _font(bold=True, color=C_FLOOR_FG), _fill(C_FLOOR_BG), _align("left")
            calc_row += 1

        bg = C_CALC_BG if idx % 2 == 0 else C_ROW_PLAIN
        formula_clean = r.formula_string.replace("\n", "  |  ")
        row_vals = [r.bar_mark, f"{r.member_type}: {r.member_name}", r.diameter_mm, r.shape_code, formula_clean, r.bend_deduction_mm]
        _set_row(ws, calc_row, row_vals, bg=bg, number_formats=[None, None, None, None, None, "#,##0"])
        ws.cell(row=calc_row, column=5).alignment = _align("left", wrap=True)
        ws.row_dimensions[calc_row].height = 30
        calc_row += 1

    if standard_module:
        calc_row += 2
        ws.merge_cells(start_row=calc_row, start_column=1, end_row=calc_row, end_column=6)
        c = ws.cell(row=calc_row, column=1, value="COMMON SHAPE REFERENCE — " + standard_module.STANDARD_NAME)
        c.font, c.fill, c.alignment = _font(bold=True, size=11, color=C_HEADER_FG), _fill(C_HEADER_BG), _align("center")
        calc_row += 1

        _set_row(ws, calc_row, ["Shape Code", "Description", "", "", "General Formula", "Notes"], bold=True, bg=C_COL_HEAD_BG, fg=C_COL_HEAD_FG)
        calc_row += 1

        for key, entry in standard_module.SHAPE_MAP.items():
            code  = entry.get("code", key)
            desc  = entry.get("description", "")
            form  = entry.get("formula", "")
            is_common = code in getattr(standard_module, "COMMON_SHAPES", [])
            bg = "FFF9C4" if is_common else C_ROW_PLAIN
            _set_row(ws, calc_row, [code, desc, "", "", form, "★ Common" if is_common else ""], bg=bg)
            ws.cell(row=calc_row, column=5).alignment = _align("left", wrap=True)
            ws.row_dimensions[calc_row].height = 24
            calc_row += 1

def _write_summary_sheet(ws, records, project_info):
    # Dynamically compute merge width from actual diameter count + fixed columns
    _all_dias = sorted(set(r.diameter_mm for r in records))
    _n_cols   = 3 + len(_all_dias) + 1  # Level, MemberType, Member + dias + Total
    _end_col  = get_column_letter(max(_n_cols, 4))
    ws.merge_cells(f"A1:{_end_col}1")
    c = ws.cell(row=1, column=1, value=f"BBS SUMMARY — {project_info.get('project_name','')}")
    c.font, c.fill, c.alignment = _font(bold=True, size=12, color=C_HEADER_FG), _fill(C_HEADER_BG), _align("center")

    row = 3
    ws.cell(row=row, column=1, value="WEIGHT SUMMARY (kg) BY LEVEL / MEMBER / DIAMETER")
    ws.cell(row=row, column=1).font = _font(bold=True, size=11)
    row += 1

    all_dias = sorted(set(r.diameter_mm for r in records))
    headers = ["Level", "Member Type", "Member"] + [f"T{d}" for d in all_dias] + ["Total (kg)"]
    _set_row(ws, row, headers, bold=True, bg=C_COL_HEAD_BG, fg=C_COL_HEAD_FG)
    row += 1

    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in records: grouped[r.floor_level][f"{r.member_type}:{r.member_name}"][r.diameter_mm] += r.total_weight_kg

    for level in sorted(grouped.keys()):
        level_total = defaultdict(float)
        for member_key in sorted(grouped[level].keys()):
            mtype, mname = member_key.split(":", 1)
            dia_weights = grouped[level][member_key]
            row_vals = [level, mtype, mname]
            row_total = 0.0
            for d in all_dias:
                w = dia_weights.get(d, 0.0)
                row_vals.append(w if w else "")
                row_total += w
                level_total[d] += w
            row_vals.append(row_total)
            _set_row(ws, row, row_vals, number_formats=[None, None, None] + ["#,##0.000"] * len(all_dias) + ["#,##0.000"])
            row += 1

        lt_vals = [level, "LEVEL TOTAL", ""] + [level_total.get(d, "") for d in all_dias]
        lt_vals.append(sum(level_total.values()))
        _set_row(ws, row, lt_vals, bold=True, bg=C_SUBTOTAL_BG, number_formats=[None, None, None] + ["#,##0.000"] * len(all_dias) + ["#,##0.000"])
        row += 2

    grand = defaultdict(float)
    for r in records: grand[r.diameter_mm] += r.total_weight_kg
    gt_vals = ["GRAND TOTAL", "", ""] + [grand.get(d, 0.0) for d in all_dias]
    gt_vals.append(sum(grand.values()))
    _set_row(ws, row, gt_vals, bold=True, bg=C_GRAND_BG, number_formats=[None, None, None] + ["#,##0.000"] * len(all_dias) + ["#,##0.000"])
    row += 2

    row += 1
    ws.cell(row=row, column=1, value="DIAMETER SUMMARY")
    ws.cell(row=row, column=1).font = _font(bold=True, size=11)
    row += 1
    _set_row(ws, row, ["Diameter", "No. of Bars", "Total Length (m)", "Total Weight (kg)"], bold=True, bg=C_COL_HEAD_BG, fg=C_COL_HEAD_FG)
    row += 1

    dia_bars   = defaultdict(int)
    dia_len    = defaultdict(float)
    dia_weight = defaultdict(float)
    for r in records:
        dia_bars[r.diameter_mm]   += r.no_of_bars
        dia_len[r.diameter_mm]    += r.total_length_mm / 1000.0
        dia_weight[r.diameter_mm] += r.total_weight_kg

    for d in sorted(dia_bars.keys()):
        _set_row(ws, row, [f"T{d}", dia_bars[d], dia_len[d], dia_weight[d]], number_formats=[None, "#,##0", "#,##0.00", "#,##0.000"])
        row += 1

    _set_row(ws, row, ["TOTAL", sum(dia_bars.values()), sum(dia_len.values()), sum(dia_weight.values())], bold=True, bg=C_GRAND_BG, number_formats=[None, "#,##0", "#,##0.00", "#,##0.000"])