# -*- coding: utf-8 -*-
"""
output/pdf_exporter.py
Exports the BBS Excel workbook to PDF.

Priority chain:
  1. Excel COM (win32com / .NET Interop) — best fidelity, requires Excel
  2. fpdf2 (pure Python) — fallback when Excel not available, requires fpdf2
  3. Graceful failure with clear install message

fpdf2 install:
  pip install fpdf2 --target "<pyRevit CPython Lib/site-packages>"
"""

import os
import sys


def export_pdf(xlsx_path, pdf_path=None):
    """
    Exports xlsx to PDF. Returns (success, message, pdf_path|None).
    Tries COM first, fpdf2 second.
    """
    if not os.path.exists(xlsx_path):
        return False, f"Excel file not found: {xlsx_path}", None

    if pdf_path is None:
        pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"

    # ── Primary: Excel COM Interop ────────────────────────────────────────────
    ok, msg, out = _export_via_com(xlsx_path, pdf_path)
    if ok:
        return True, msg, out

    # ── Fallback: fpdf2 ───────────────────────────────────────────────────────
    ok2, msg2, out2 = _export_via_fpdf2(xlsx_path, pdf_path)
    if ok2:
        return True, msg2, out2

    # Both failed
    return False, (
        "PDF export failed.\n"
        "Primary (Excel COM): " + msg + "\n"
        "Fallback (fpdf2): " + msg2 + "\n\n"
        "To enable fpdf2 fallback install it with:\n"
        "  pip install fpdf2 --target \"<pyRevit CPython Lib/site-packages>\""
    ), None


def _export_via_com(xlsx_path, pdf_path):
    """Export via Excel COM Interop (.NET). Requires Excel installed."""
    try:
        import System
        from System import Type, Activator, Reflection

        excel_type = Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return False, "Excel not installed", None

        f_set = Reflection.BindingFlags.SetProperty
        f_inv = Reflection.BindingFlags.InvokeMethod
        f_get = Reflection.BindingFlags.GetProperty

        excel_app = Activator.CreateInstance(excel_type)
        excel_type.InvokeMember("Visible",       f_set, None, excel_app, System.Array[object]([False]))
        excel_type.InvokeMember("DisplayAlerts", f_set, None, excel_app, System.Array[object]([False]))

        workbooks = excel_type.InvokeMember("Workbooks", f_get, None, excel_app, None)
        workbook  = workbooks.GetType().InvokeMember(
            "Open", f_inv, None, workbooks, System.Array[object]([xlsx_path]))

        workbook.GetType().InvokeMember(
            "ExportAsFixedFormat", f_inv, None, workbook,
            System.Array[object]([0, pdf_path]))  # 0 = xlTypePDF

        workbook.GetType().InvokeMember("Close", f_inv, None, workbook,
                                        System.Array[object]([False]))
        excel_type.InvokeMember("Quit", f_inv, None, excel_app, None)
        try:
            System.Runtime.InteropServices.Marshal.ReleaseComObject(excel_app)
        except Exception:
            pass

        return True, f"PDF saved via Excel COM: {pdf_path}", pdf_path

    except Exception as ex:
        return False, str(ex), None


def _export_via_fpdf2(xlsx_path, pdf_path):
    """
    Export via fpdf2 — pure Python fallback.
    Reads the BBS sheet from the xlsx and creates a formatted PDF table.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        return False, "fpdf2 not installed", None

    try:
        import openpyxl
        wb  = openpyxl.load_workbook(xlsx_path, data_only=True)

        pdf = FPDF(orientation="L", unit="mm", format="A3")
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.set_margins(10, 10, 10)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            pdf.add_page()

            # Sheet title
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_fill_color(20, 20, 20)      # Charcoal Black
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 8, f"AnonGee BIM Tools  |  {sheet_name}", fill=True, ln=True)
            pdf.ln(2)

            # Column widths — auto from first row
            max_col = ws.max_column
            col_w   = min(35, 267 / max(max_col, 1))

            pdf.set_font("Helvetica", "B", 7)
            pdf.set_fill_color(20, 20, 20)
            pdf.set_text_color(255, 255, 255)

            row_num = 0
            for row in ws.iter_rows(values_only=True):
                if all(c is None for c in row):
                    pdf.ln(2)
                    continue
                row_num += 1
                is_header = row_num == 1

                if is_header:
                    pdf.set_font("Helvetica", "B", 7)
                    pdf.set_fill_color(20, 20, 20)
                    pdf.set_text_color(255, 255, 255)
                    fill = True
                else:
                    pdf.set_font("Helvetica", "", 7)
                    if row_num % 2 == 0:
                        pdf.set_fill_color(244, 244, 246)  # Off White
                    else:
                        pdf.set_fill_color(255, 255, 255)
                    pdf.set_text_color(20, 20, 20)
                    fill = True

                for cell in row:
                    val = str(cell) if cell is not None else ""
                    # Truncate long values
                    if len(val) > 18:
                        val = val[:17] + "."
                    pdf.cell(col_w, 5, val, border=1, fill=fill)
                pdf.ln()

        pdf.output(pdf_path)
        return True, f"PDF saved via fpdf2: {pdf_path}", pdf_path

    except Exception as ex:
        return False, f"fpdf2 export error: {ex}", None
